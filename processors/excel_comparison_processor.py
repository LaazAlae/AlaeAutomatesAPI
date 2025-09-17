from flask import Blueprint, request, jsonify, send_file
import pandas as pd
import openpyxl
import json
import logging
import os
import tempfile
import hashlib
import re
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from werkzeug.utils import secure_filename
from difflib import SequenceMatcher
import warnings
warnings.filterwarnings('ignore')

excel_comparison_bp = Blueprint('excel_comparison', __name__)

# Configuration
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
REQUIRED_COLUMN_NAMES = ['GroupName', 'CorpName', 'Amount_To_Apply', 'ReceiptType', 'ReceiptNumber', 'RecBatchName', 'ReceiptCreateDate', 'ReceiptsID', 'CorpID', 'GroupID', 'RecBatchID', 'PostDate', 'SourceName', 'Notes', 'Date Last Change', 'User Last Change']
COMPARISON_COLUMNS = ['GroupName', 'CorpName', 'ReceiptType', 'ReceiptNumber', 'RecBatchName', 'ReceiptCreateDate', 'ReceiptsID', 'CorpID', 'GroupID', 'RecBatchID']

# Create a persistent directory for processed files
PROCESSED_FILES_DIR = os.path.join(tempfile.gettempdir(), 'excel_comparison_files')
os.makedirs(PROCESSED_FILES_DIR, exist_ok=True)

def cleanup_old_files():
    """Clean up files older than 1 hour to prevent disk space issues"""
    try:
        current_time = datetime.now()
        for filename in os.listdir(PROCESSED_FILES_DIR):
            if filename.endswith('_comparison.xlsx'):
                file_path = os.path.join(PROCESSED_FILES_DIR, filename)
                # Get file modification time
                file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
                # Remove files older than 1 hour
                if (current_time - file_mtime).total_seconds() > 3600:  # 1 hour
                    os.remove(file_path)
                    logging.info(f"Cleaned up old file: {filename}")
    except Exception as e:
        logging.error(f"Cleanup error: {e}")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def normalize_text_for_comparison(text_to_normalize):
    return str(text_to_normalize).lower().replace(' ', '').replace('_', '').replace('-', '').replace('(', '').replace(')', '').strip()

def calculate_header_similarity_score(required_column_name, excel_cell_text):
    """Ultra-flexible matching for header variations"""
    normalized_required_name = normalize_text_for_comparison(required_column_name)
    normalized_cell_text = normalize_text_for_comparison(excel_cell_text)

    # Exact normalized match gets perfect score
    if normalized_required_name == normalized_cell_text: return 100

    # Dictionary mapping words to their common variations and synonyms
    word_variations_dictionary = {
        'create': ['creation', 'created', 'date'],
        'corp': ['corporation', 'company', 'corporate'],
        'group': ['grp'],
        'receipt': ['rec', 'rcpt'],
        'receipts': ['receipt'],  # ReceiptsID -> Receipt ID
        'batch': ['btch'],
        'amount': ['amt', 'total'],
        'number': ['num', 'no', '#'],
        'source': ['src'],
        'date': ['dt', 'time'],
        'change': ['changed', 'modify', 'modified'],
        'last': ['final'],
        'user': ['usr', 'person'],
        'name': [],  # GroupName could just be "Group"
        'apply': ['application', 'applied']
    }

    # Split column names into individual words for comparison
    required_column_words = normalized_required_name.replace('_', ' ').split()
    excel_cell_words = normalized_cell_text.replace('_', ' ').split()

    # Handle specific common patterns that are known to match
    if required_column_name == 'ReceiptsID' and 'receipt' in normalized_cell_text and 'id' in normalized_cell_text:
        return 95  # "Receipt ID" -> "ReceiptsID"
    if required_column_name == 'CorpID' and 'corporation' in normalized_cell_text and 'id' in normalized_cell_text:
        return 95  # "Corporation ID" -> "CorpID"
    if required_column_name == 'ReceiptCreateDate' and 'receipt' in normalized_cell_text and ('creation' in normalized_cell_text or 'create' in normalized_cell_text) and 'date' in normalized_cell_text:
        return 95  # "Receipt Creation Date" -> "ReceiptCreateDate"

    # Count how many words match between required name and excel cell
    matching_words_count = 0
    total_required_words = len(required_column_words)

    for required_word in required_column_words:
        word_was_found = False
        # Look for exact word match first
        if required_word in excel_cell_words:
            matching_words_count += 1
            word_was_found = True
        else:
            # Check if any variations of this word exist
            word_variations_list = word_variations_dictionary.get(required_word, [])
            for variation in word_variations_list:
                if variation in normalized_cell_text:
                    matching_words_count += 1
                    word_was_found = True
                    break
            # Check if word is partially contained in any excel cell word
            if not word_was_found:
                for excel_word in excel_cell_words:
                    if len(required_word) >= 3 and len(excel_word) >= 3:
                        if required_word in excel_word or excel_word in required_word:
                            matching_words_count += 0.8  # Partial match gets lower score
                            word_was_found = True
                            break

        # Handle cases where "Name" suffix might be missing from excel headers
        if not word_was_found and required_word == 'name' and any(word in ['group', 'corp'] for word in required_column_words):
            matching_words_count += 0.5  # Allow missing "name" suffix if we have group/corp

    # Calculate final similarity score based on word match percentage
    word_match_percentage = matching_words_count / total_required_words if total_required_words > 0 else 0
    if word_match_percentage >= 0.7:  # 70% of words match - high confidence
        return int(75 + (word_match_percentage * 25))
    elif word_match_percentage >= 0.4:  # 40% of words match - medium confidence
        return int(40 + (word_match_percentage * 35))
    else:
        return 0  # Too few words match - not a good match

def load_excel_file(excel_file_path):
    """Load and process Excel file with header detection"""
    try:
        logging.info(f"Processing: {os.path.basename(excel_file_path)}")

        excel_workbook = openpyxl.load_workbook(excel_file_path, read_only=False, data_only=False)
        active_worksheet = excel_workbook.active
        all_excel_rows = list(active_worksheet.values)
        excel_workbook.close()

        successfully_found_columns = {}
        for required_column_name in REQUIRED_COLUMN_NAMES:
            best_match_for_this_column = {'score': 0, 'data': []}

            for row_index, excel_row in enumerate(all_excel_rows[:100]):
                if not excel_row: continue
                for column_index, excel_cell in enumerate(excel_row):
                    if excel_cell is None: continue
                    excel_cell_text = str(excel_cell).strip()
                    if len(excel_cell_text) < 2: continue  # Skip single characters

                    similarity_score = calculate_header_similarity_score(required_column_name, excel_cell_text)

                    if similarity_score > best_match_for_this_column['score']:
                        column_data = [all_excel_rows[i][column_index] for i in range(row_index+1, len(all_excel_rows)) if column_index < len(all_excel_rows[i]) and all_excel_rows[i][column_index] is not None and str(all_excel_rows[i][column_index]).strip()]
                        if column_data:
                            best_match_for_this_column = {'score': similarity_score, 'data': column_data}

            if best_match_for_this_column['score'] > 0:
                successfully_found_columns[required_column_name] = best_match_for_this_column

        if len(successfully_found_columns) < 5:
            return {'error': 'Too few columns found'}

        # Get the maximum data length and create unified DataFrame
        maximum_data_length = max(len(column_info['data']) for column_info in successfully_found_columns.values())

        # Create DataFrame with normalized column lengths
        unified_dataframe = pd.DataFrame({
            column_name: column_info['data'] + [None] * (maximum_data_length - len(column_info['data']))
            for column_name, column_info in successfully_found_columns.items()
        })

        # Format date columns using pandas
        for date_column_name in ['ReceiptCreateDate', 'PostDate', 'Date Last Change']:
            if date_column_name in unified_dataframe.columns:
                unified_dataframe[date_column_name] = pd.to_datetime(unified_dataframe[date_column_name], errors='coerce')

        # Format receipt numbers using pandas apply
        if 'ReceiptNumber' in unified_dataframe.columns:
            unified_dataframe['ReceiptNumber'] = unified_dataframe['ReceiptNumber'].apply(
                lambda x: int(x) if pd.notna(x) and str(x).strip().lstrip('-').replace('.', '').isdigit() and not any(c.isalpha() for c in str(x)) else x
            )

        logging.info(f"Success! Processed {os.path.basename(excel_file_path)} with {len(successfully_found_columns)} columns")
        return {'dataframe': unified_dataframe}

    except Exception as e:
        logging.error(f"Excel loading error: {str(e)}")
        return {'error': str(e)}

def hash_row(row):
    """Create hash for row comparison"""
    return hashlib.md5('|'.join(
        str(row.get(c, '')).strip().lower() if not isinstance(row.get(c), (int, float)) or not float(row.get(c)).is_integer()
        else str(int(row.get(c))) for c in COMPARISON_COLUMNS
    ).encode()).hexdigest()

def filter_date(df, cutoff_month_year):
    """Filter out records from specified month/year and after"""
    if 'ReceiptCreateDate' not in df.columns:
        return df, []

    try:
        month, year = cutoff_month_year.split('/')
        cutoff_date = datetime(int(year), int(month), 1)
        mask = pd.isna(df['ReceiptCreateDate']) | (df['ReceiptCreateDate'] < cutoff_date)
        filtered_df = df[mask].reset_index(drop=True)
        removed_records = df[~mask].to_dict('records')
        return filtered_df, removed_records
    except:
        return df, []

def fuzzy_industry_match(r1, r2):
    """Advanced fuzzy matching for industry-specific records"""
    def normalize_company(name):
        if not name: return ''
        name = str(name).strip().lower()
        # Remove common business suffixes
        name = re.sub(r'\s*,?\s*(llc|inc\.?|corp\.?|ltd\.?|lp|llp|p\.?c\.?)\s*$', '', name)
        # Replace & with 'and'
        name = re.sub(r'[&]', 'and', name)
        # Remove special characters
        name = re.sub(r'[^\w\s]', ' ', name)
        # Normalize whitespace
        name = re.sub(r'\s+', ' ', name).strip()
        return name

    def normalize_person(name):
        if not name: return ''
        name = str(name).strip().lower()
        # Remove legal suffixes
        name = re.sub(r'\s*,?\s*(esq\.?(\(.*?\))?|legal\s+assistant|\(.*?\))\s*$', '', name)
        # Remove special characters
        name = re.sub(r'[^\w\s]', ' ', name)
        # Normalize whitespace
        name = re.sub(r'\s+', ' ', name).strip()
        return name

    def similarity(s1, s2):
        if s1 == s2: return 1.0
        if not s1 or not s2: return 0.0
        return SequenceMatcher(None, s1, s2).ratio()

    exact_match_fields = {'ReceiptCreateDate', 'RecBatchName', 'ReceiptType', 'ReceiptsID', 'CorpID', 'GroupID', 'RecBatchID'}
    matches = 0
    scores = {}
    differences = []

    for field in COMPARISON_COLUMNS:
        v1, v2 = str(r1.get(field, '')).strip(), str(r2.get(field, '')).strip()

        if v1.lower() == v2.lower():
            matches += 1
            scores[field] = 1.0
        elif field == 'GroupName':
            n1, n2 = normalize_company(v1), normalize_company(v2)
            score = similarity(n1, n2)
            # Boost score if one name contains the other
            if n1 and n2 and (n1 in n2 or n2 in n1):
                score = max(score, 0.90)
            scores[field] = score
            if score >= 0.85:
                matches += 1
            else:
                differences.append(field)
        elif field == 'CorpName':
            n1, n2 = normalize_person(v1), normalize_person(v2)
            score = similarity(n1, n2)
            scores[field] = score
            if score >= 0.90:
                matches += 1
            else:
                differences.append(field)
        elif field == 'ReceiptNumber':
            # Normalize receipt numbers
            n1 = re.sub(r'[-\s]', '', v1.upper())
            n2 = re.sub(r'[-\s]', '', v2.upper())
            score = similarity(n1, n2)
            scores[field] = score
            if score >= 0.80:
                matches += 1
            else:
                differences.append(field)
        elif field in exact_match_fields:
            # Handle numeric fields that might be integers
            try:
                if isinstance(r1.get(field), (int, float)) and float(r1.get(field)).is_integer():
                    nv1 = str(int(float(r1.get(field))))
                else:
                    nv1 = v1
                if isinstance(r2.get(field), (int, float)) and float(r2.get(field)).is_integer():
                    nv2 = str(int(float(r2.get(field))))
                else:
                    nv2 = v2
            except:
                nv1, nv2 = v1, v2

            if nv1.lower() == nv2.lower():
                scores[field] = 1.0
                matches += 1
            else:
                scores[field] = 0.0
                differences.append(field)
        else:
            score = similarity(v1.lower(), v2.lower())
            scores[field] = score
            if score >= 0.95:
                matches += 1
            else:
                differences.append(field)

    # Determine if it's a match based on multiple criteria
    match_ratio = matches / len(COMPARISON_COLUMNS)
    avg_score = sum(scores.values()) / len(COMPARISON_COLUMNS)
    is_match = match_ratio >= 0.80 and avg_score >= 0.85 and len(differences) <= 2

    return is_match, avg_score, differences, scores

def compare_dataframes(df1, df2):
    """Compare two dataframes and find differences"""
    # Create hash maps for quick comparison
    hash1 = {hash_row(row): idx for idx, row in df1.iterrows()}
    hash2 = {hash_row(row): idx for idx, row in df2.iterrows()}

    # Find exact matches and differences
    common_hashes = set(hash1.keys()) & set(hash2.keys())
    only_in_df1 = set(hash1.keys()) - set(hash2.keys())
    only_in_df2 = set(hash2.keys()) - set(hash1.keys())

    # Fuzzy matching for unmatched records
    fuzzy_matches = []
    remaining_df1 = set(only_in_df1)
    remaining_df2 = set(only_in_df2)

    logging.info(f"Running fuzzy matching on {len(only_in_df1)} vs {len(only_in_df2)} unmatched records...")

    for hash1_key in list(only_in_df1):
        if hash1_key not in remaining_df1:
            continue

        row1 = df1.iloc[hash1[hash1_key]]
        best_hash = None
        best_confidence = 0

        # Limit search to prevent timeout
        search_limit = min(500, len(remaining_df2))
        for hash2_key in list(remaining_df2)[:search_limit]:
            row2 = df2.iloc[hash2[hash2_key]]
            is_match, confidence, _, _ = fuzzy_industry_match(row1, row2)

            if is_match and confidence > best_confidence:
                best_hash = hash2_key
                best_confidence = confidence

        if best_hash:
            row2 = df2.iloc[hash2[best_hash]]
            try:
                # Compare amounts
                amount1 = float(str(row1.get('Amount_To_Apply', 0)).replace('$', '').replace(',', ''))
                amount2 = float(str(row2.get('Amount_To_Apply', 0)).replace('$', '').replace(',', ''))

                if abs(amount1 - amount2) > 0.001:  # Different amounts
                    fuzzy_row = row1.copy()
                    fuzzy_row['Amount_To_Apply_M1'] = amount1
                    fuzzy_row['Amount_To_Apply_M2'] = amount2
                    fuzzy_row['Fuzzy_Match_Confidence'] = f'{best_confidence:.3f}'
                    del fuzzy_row['Amount_To_Apply']
                    fuzzy_matches.append(fuzzy_row)
            except:
                pass

            remaining_df1.discard(hash1_key)
            remaining_df2.discard(best_hash)

    # Check amount differences in exact matches
    amount_differences = []
    for common_hash in common_hashes:
        row1 = df1.iloc[hash1[common_hash]]
        row2 = df2.iloc[hash2[common_hash]]

        try:
            amount1 = float(str(row1.get('Amount_To_Apply', 0)).replace('$', '').replace(',', ''))
            amount2 = float(str(row2.get('Amount_To_Apply', 0)).replace('$', '').replace(',', ''))

            if abs(amount1 - amount2) > 0.001:  # Different amounts
                diff_row = row1.copy()
                diff_row['Amount_To_Apply_M1'] = amount1
                diff_row['Amount_To_Apply_M2'] = amount2
                del diff_row['Amount_To_Apply']
                amount_differences.append(diff_row)
        except:
            pass

    # Get records only in each month
    month1_only = df1.iloc[[hash1[h] for h in remaining_df1]].reset_index(drop=True)
    month2_only = df2.iloc[[hash2[h] for h in remaining_df2]].reset_index(drop=True)

    logging.info(f"Fuzzy matching found {len(fuzzy_matches)} additional matches!")

    # Combine amount differences and fuzzy matches
    all_matches = pd.DataFrame(amount_differences + fuzzy_matches)

    return all_matches, month1_only, month2_only

def filter_data(df):
    """Filter out incomplete or summary rows"""
    if len(df) == 0:
        return df

    key_fields = [f for f in ['GroupName', 'CorpName', 'ReceiptType', 'ReceiptNumber'] if f in df.columns]
    if not key_fields:
        return df

    filtered_rows = []
    for _, row in df.iterrows():
        # Count populated key fields
        populated_count = sum(1 for f in key_fields
                            if pd.notna(row.get(f)) and str(row.get(f)).strip() and str(row.get(f)).strip().lower() != 'nan')

        # Check for summary row keywords
        group_name = str(row.get('GroupName', '')).strip().lower()
        is_summary = any(word in group_name for word in ['total', 'sum', 'summary', 'subtotal'])

        # Keep row if it has enough populated fields and isn't a summary
        if populated_count >= len(key_fields) * 0.5 and not is_summary:
            filtered_rows.append(row)

    return pd.DataFrame(filtered_rows).reset_index(drop=True) if filtered_rows else pd.DataFrame()

def create_comparison_excel(matches, month1_only, month2_only, output_path):
    """Create Excel file with comparison results"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Comparison Results"

    # Define styles
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    current_row = 1
    column_widths = {}

    # Process each section
    sections = [
        ("AMOUNT CHANGES", matches, "Records with same data but different amounts"),
        ("MONTH 1 ONLY", month1_only, "Records only in Month 1"),
        ("MONTH 2 ONLY", month2_only, "Records only in Month 2")
    ]

    for section_name, data, description in sections:
        # Add description
        cell = worksheet.cell(current_row, 1, f"{description} | Total: {len(data)}")
        cell.font = Font(name="Calibri", size=10, italic=True)
        worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=20)
        current_row += 1

        if len(data) > 0:
            # Add title
            title_cell = worksheet.cell(current_row, 1, f"TABLE: {section_name}")
            title_cell.font = title_font
            title_cell.fill = title_fill

            # Get columns
            columns = list(data.columns)
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
            current_row += 1

            # Add headers
            for col_idx, column in enumerate(columns, 1):
                cell = worksheet.cell(current_row, col_idx, column.replace('_', ' '))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                column_widths[cell.column_letter] = max(column_widths.get(cell.column_letter, 8), len(column.replace('_', ' ')) + 2)
            current_row += 1

            # Add data
            for _, row in data.iterrows():
                for col_idx, column in enumerate(columns, 1):
                    cell = worksheet.cell(current_row, col_idx, row[column])
                    cell.font = data_font

                    # Apply special formatting
                    if 'Amount' in column and (column.endswith('Apply') or column.endswith('M1') or column.endswith('M2')):
                        cell.number_format = '"$"#,##0.00_);\\("$"#,##0.00\\)'
                    elif 'Date' in column and isinstance(row[column], datetime):
                        cell.number_format = 'mm-dd-yy'

                    # Update column width
                    value_length = len(str(cell.value)) + 2 if cell.value else 8
                    max_width = 25 if 'Notes' in column else 50
                    column_widths[cell.column_letter] = min(max_width, max(column_widths.get(cell.column_letter, 8), value_length))

                current_row += 1
        else:
            # Empty section
            title_cell = worksheet.cell(current_row, 1, f"TABLE: {section_name}")
            title_cell.font = title_font
            title_cell.fill = title_fill
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
            current_row += 1

            cell = worksheet.cell(current_row, 1, "No records found")
            cell.font = Font(name="Calibri", size=11, italic=True, color="999999")
            current_row += 1

        current_row += 2  # Add spacing between sections

    # Set column widths
    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    # Freeze top row
    worksheet.freeze_panes = 'A2'

    # Save workbook
    workbook.save(output_path)

@excel_comparison_bp.route('/', methods=['GET'])
def get_service_info():
    """Get Excel Comparison service information"""
    return jsonify({
        'service': 'Excel Comparison API',
        'description': 'Compare two Excel files month-to-month and identify differences in amounts and records',
        'usage': 'POST with month1_file, month2_file, and cutoff_month parameters',
        'supported_formats': ['xlsx', 'xls'],
        'required_inputs': {
            'month1_file': 'Excel file for first month',
            'month2_file': 'Excel file for second month',
            'cutoff_month': 'Month/Year to filter (e.g., "08/2025")'
        },
        'outputs': {
            'amount_changes': 'Records with same data but different amounts',
            'month1_only': 'Records only found in month 1',
            'month2_only': 'Records only found in month 2'
        }
    })

@excel_comparison_bp.route('/process', methods=['POST'])
def process_comparison():
    """Process Excel comparison with three inputs"""
    logging.info("Excel comparison processing request received")

    # Clean up old files before processing
    cleanup_old_files()

    try:
        # Check if all required files and data are uploaded
        if 'month1_file' not in request.files:
            return jsonify({'success': False, 'error': 'No month 1 file uploaded'}), 400
        if 'month2_file' not in request.files:
            return jsonify({'success': False, 'error': 'No month 2 file uploaded'}), 400
        if 'cutoff_month' not in request.form:
            return jsonify({'success': False, 'error': 'No cutoff month provided'}), 400

        month1_file = request.files['month1_file']
        month2_file = request.files['month2_file']
        cutoff_month = request.form['cutoff_month'].strip()

        # Validate files
        if month1_file.filename == '' or month2_file.filename == '':
            return jsonify({'success': False, 'error': 'No files selected'}), 400

        if not allowed_file(month1_file.filename) or not allowed_file(month2_file.filename):
            return jsonify({'success': False, 'error': 'Invalid file type. Please upload Excel files (.xlsx or .xls)'}), 400

        # Validate cutoff month format
        try:
            month, year = cutoff_month.split('/')
            int(month), int(year)
        except:
            return jsonify({'success': False, 'error': 'Invalid cutoff month format. Use MM/YYYY (e.g., "08/2025")'}), 400

        # Save uploaded files temporarily
        temp_files = []
        try:
            # Save month 1 file
            filename1 = secure_filename(month1_file.filename)
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file1:
                month1_file.save(tmp_file1.name)
                temp_month1_path = tmp_file1.name
                temp_files.append(temp_month1_path)

            # Save month 2 file
            filename2 = secure_filename(month2_file.filename)
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file2:
                month2_file.save(tmp_file2.name)
                temp_month2_path = tmp_file2.name
                temp_files.append(temp_month2_path)

            # Load and process both Excel files
            result1 = load_excel_file(temp_month1_path)
            result2 = load_excel_file(temp_month2_path)

            if 'error' in result1:
                return jsonify({'success': False, 'error': f'Error processing month 1 file: {result1["error"]}'}), 400
            if 'error' in result2:
                return jsonify({'success': False, 'error': f'Error processing month 2 file: {result2["error"]}'}), 400

            df1 = result1['dataframe']
            df2 = result2['dataframe']

            # Add original line numbers for tracking
            df1['Original_Line'] = range(2, len(df1) + 2)
            df2['Original_Line'] = range(2, len(df2) + 2)

            # Filter data to remove incomplete/summary rows
            df1_filtered = filter_data(df1)
            df2_filtered = filter_data(df2)

            # Apply date filter to month 2
            df2_date_filtered, removed_records = filter_date(df2_filtered, cutoff_month)

            # Perform comparison
            matches, month1_only, month2_only = compare_dataframes(df1_filtered, df2_date_filtered)

            # Generate unique file ID and create output file
            import uuid
            file_id = str(uuid.uuid4())
            output_file_path = os.path.join(PROCESSED_FILES_DIR, f'{file_id}_comparison.xlsx')

            # Create Excel file with results
            create_comparison_excel(matches, month1_only, month2_only, output_file_path)

            return jsonify({
                'success': True,
                'message': 'Excel comparison completed successfully',
                'file_id': file_id,
                'summary': {
                    'month1_total': len(df1_filtered),
                    'month2_total': len(df2_filtered),
                    'month2_after_date_filter': len(df2_date_filtered),
                    'removed_by_date_filter': len(removed_records),
                    'amount_changes': len(matches),
                    'month1_only': len(month1_only),
                    'month2_only': len(month2_only)
                },
                'cutoff_month': cutoff_month,
                'download_ready': True
            })

        finally:
            # Clean up temporary input files
            for temp_path in temp_files:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)

    except Exception as e:
        logging.error(f"Excel comparison processing error: {str(e)}")
        return jsonify({'success': False, 'error': f'Processing failed: {str(e)}'}), 500

@excel_comparison_bp.route('/download/<file_id>', methods=['GET'])
def download_comparison_file(file_id):
    """Download the comparison results Excel file using file ID"""
    try:
        # Validate file_id format (UUID)
        import uuid
        try:
            uuid.UUID(file_id)
        except ValueError:
            return jsonify({'error': 'Invalid file ID format'}), 400

        # Construct file path based on file ID
        file_path = os.path.join(PROCESSED_FILES_DIR, f'{file_id}_comparison.xlsx')

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found or expired. Files are automatically cleaned up after some time.'}), 404

        # Generate a filename with timestamp
        current_date = datetime.now()
        filename = f'excel_comparison_results_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.error(f"Download error: {str(e)}")
        return jsonify({'error': f'Download failed: {str(e)}'}), 500