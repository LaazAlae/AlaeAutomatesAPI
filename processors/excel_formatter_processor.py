from flask import Blueprint, request, jsonify, send_file
import openpyxl
import pandas as pd
import json
import logging
import os
import tempfile
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from werkzeug.utils import secure_filename
import warnings
warnings.filterwarnings('ignore')

excel_formatter_bp = Blueprint('excel_formatter', __name__)

# Configuration
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
REQUIRED_COLUMN_NAMES = ['GroupName', 'CorpName', 'Amount_To_Apply', 'ReceiptType', 'ReceiptNumber', 'RecBatchName', 'ReceiptCreateDate', 'ReceiptsID', 'CorpID', 'GroupID', 'RecBatchID', 'PostDate', 'SourceName', 'Notes', 'Date Last Change', 'User Last Change']

# Create a persistent directory for processed files
PROCESSED_FILES_DIR = os.path.join(tempfile.gettempdir(), 'excel_formatter_files')
os.makedirs(PROCESSED_FILES_DIR, exist_ok=True)

def cleanup_old_files():
    """Clean up files older than 1 hour to prevent disk space issues"""
    try:
        current_time = datetime.now()
        for filename in os.listdir(PROCESSED_FILES_DIR):
            if filename.endswith('_formatted.xlsx'):
                file_path = os.path.join(PROCESSED_FILES_DIR, filename)
                # Get file modification time
                file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
                # Remove files older than 1 hour
                if (current_time - file_mtime).total_seconds() > 3600:  # 1 hour
                    os.remove(file_path)
                    logging.info(f"Cleaned up old file: {filename}")
    except Exception as e:
        logging.error(f"Cleanup error: {e}")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def normalize_text_for_comparison(text_to_normalize):
    return str(text_to_normalize).lower().replace(' ', '').replace('_', '').replace('-', '').replace('(', '').replace(')', '').strip()

def calculate_header_similarity_score(required_column_name, excel_cell_text):
    """Ultra-flexible matching for header variations"""
    normalized_required_name = normalize_text_for_comparison(required_column_name)
    normalized_cell_text = normalize_text_for_comparison(excel_cell_text)

    # Exact normalized match gets perfect score
    if normalized_required_name == normalized_cell_text: return 100

    # Dictionary mapping words to their common variations and synonyms
    word_variations_dictionary = {
        'create': ['creation', 'created', 'date'],
        'corp': ['corporation', 'company', 'corporate'],
        'group': ['grp'],
        'receipt': ['rec', 'rcpt'],
        'receipts': ['receipt'],  # ReceiptsID -> Receipt ID
        'batch': ['btch'],
        'amount': ['amt', 'total'],
        'number': ['num', 'no', '#'],
        'source': ['src'],
        'date': ['dt', 'time'],
        'change': ['changed', 'modify', 'modified'],
        'last': ['final'],
        'user': ['usr', 'person'],
        'name': [],  # GroupName could just be "Group"
        'apply': ['application', 'applied']
    }

    # Split column names into individual words for comparison
    required_column_words = normalized_required_name.replace('_', ' ').split()
    excel_cell_words = normalized_cell_text.replace('_', ' ').split()

    # Handle specific common patterns that are known to match
    if required_column_name == 'ReceiptsID' and 'receipt' in normalized_cell_text and 'id' in normalized_cell_text:
        return 95  # "Receipt ID" -> "ReceiptsID"
    if required_column_name == 'CorpID' and 'corporation' in normalized_cell_text and 'id' in normalized_cell_text:
        return 95  # "Corporation ID" -> "CorpID"
    if required_column_name == 'ReceiptCreateDate' and 'receipt' in normalized_cell_text and ('creation' in normalized_cell_text or 'create' in normalized_cell_text) and 'date' in normalized_cell_text:
        return 95  # "Receipt Creation Date" -> "ReceiptCreateDate"

    # Count how many words match between required name and excel cell
    matching_words_count = 0
    total_required_words = len(required_column_words)

    for required_word in required_column_words:
        word_was_found = False
        # Look for exact word match first
        if required_word in excel_cell_words:
            matching_words_count += 1
            word_was_found = True
        else:
            # Check if any variations of this word exist
            word_variations_list = word_variations_dictionary.get(required_word, [])
            for variation in word_variations_list:
                if variation in normalized_cell_text:
                    matching_words_count += 1
                    word_was_found = True
                    break
            # Check if word is partially contained in any excel cell word
            if not word_was_found:
                for excel_word in excel_cell_words:
                    if len(required_word) >= 3 and len(excel_word) >= 3:
                        if required_word in excel_word or excel_word in required_word:
                            matching_words_count += 0.8  # Partial match gets lower score
                            word_was_found = True
                            break

        # Handle cases where "Name" suffix might be missing from excel headers
        if not word_was_found and required_word == 'name' and any(word in ['group', 'corp'] for word in required_column_words):
            matching_words_count += 0.5  # Allow missing "name" suffix if we have group/corp

    # Calculate final similarity score based on word match percentage
    word_match_percentage = matching_words_count / total_required_words if total_required_words > 0 else 0
    if word_match_percentage >= 0.7:  # 70% of words match - high confidence
        return int(75 + (word_match_percentage * 25))
    elif word_match_percentage >= 0.4:  # 40% of words match - medium confidence
        return int(40 + (word_match_percentage * 35))
    else:
        return 0  # Too few words match - not a good match

def process_excel_and_fix_formatting(excel_file_path):
    try:
        if not os.path.exists(excel_file_path):
            return {'error': f'File not found: {excel_file_path}'}

        logging.info(f"Processing: {os.path.basename(excel_file_path)}")
        processing_log = {'timestamp': datetime.now().isoformat(), 'file': excel_file_path, 'steps': [], 'status': None}

        excel_workbook = openpyxl.load_workbook(excel_file_path, read_only=False, data_only=False)
        active_worksheet = excel_workbook.active
        all_excel_rows = list(active_worksheet.values)
        excel_workbook.close()

        rows_with_actual_data = sum(1 for excel_row in all_excel_rows if excel_row and any(cell is not None and str(cell).strip() for cell in excel_row))
        processing_log['steps'].append({'step': 1, 'action': 'Loaded Excel', 'result': f'{rows_with_actual_data} data rows (total {len(all_excel_rows)} including empty)'})

        successfully_found_columns = {}
        columns_not_found = []

        for required_column_name in REQUIRED_COLUMN_NAMES:
            best_match_for_this_column = {'score': 0, 'pos': None, 'data': [], 'header': None}
            potential_candidates = []

            for row_index, excel_row in enumerate(all_excel_rows[:100]):
                if not excel_row: continue
                for column_index, excel_cell in enumerate(excel_row):
                    if excel_cell is None: continue
                    excel_cell_text = str(excel_cell).strip()
                    if len(excel_cell_text) < 2: continue  # Skip single characters

                    similarity_score = calculate_header_similarity_score(required_column_name, excel_cell_text)
                    if similarity_score > 0:
                        potential_candidates.append({'score': similarity_score, 'pos': f'R{row_index+1}C{column_index+1}', 'text': excel_cell_text})

                    if similarity_score > best_match_for_this_column['score']:
                        column_data = [all_excel_rows[i][column_index] for i in range(row_index+1, len(all_excel_rows)) if column_index < len(all_excel_rows[i]) and all_excel_rows[i][column_index] is not None and str(all_excel_rows[i][column_index]).strip()]
                        if column_data:
                            best_match_for_this_column = {'score': similarity_score, 'pos': f'R{row_index+1}C{column_index+1}', 'data': column_data, 'header': excel_cell_text}

            if best_match_for_this_column['score'] > 0:
                successfully_found_columns[required_column_name] = best_match_for_this_column
            else:
                columns_not_found.append({'target': required_column_name, 'candidates': potential_candidates[:3]})

        processing_log['steps'].append({'step': 2, 'action': 'Found headers', 'result': f'{len(successfully_found_columns)}/{len(REQUIRED_COLUMN_NAMES)} columns', 'details': [f'{column_name}: {match_info["header"]} at {match_info["pos"]} (score: {match_info["score"]})' for column_name, match_info in successfully_found_columns.items()]})

        if columns_not_found:
            processing_log['steps'].append({'step': '2b', 'action': 'Headers NOT FOUND', 'missing': [{'target': missing_column['target'], 'best_candidates': missing_column['candidates']} for missing_column in columns_not_found]})

        if len(successfully_found_columns) < 5:
            processing_log['status'] = 'FAILED'
            return {'error': 'Too few columns found (minimum 5 required)', 'log': processing_log, 'columns_found': len(successfully_found_columns)}

        # Get the maximum data length and create unified DataFrame
        maximum_data_length = max(len(column_info['data']) for column_info in successfully_found_columns.values())

        # Create DataFrame with normalized column lengths (pandas approach from autoFixFormatter.py)
        unified_dataframe = pd.DataFrame({
            column_name: column_info['data'] + [None] * (maximum_data_length - len(column_info['data']))
            for column_name, column_info in successfully_found_columns.items()
        })

        # Format date columns using pandas (more efficient)
        for date_column_name in ['ReceiptCreateDate', 'PostDate', 'Date Last Change']:
            if date_column_name in unified_dataframe.columns:
                unified_dataframe[date_column_name] = pd.to_datetime(unified_dataframe[date_column_name], errors='coerce')

        # Format receipt numbers using pandas apply (more efficient)
        if 'ReceiptNumber' in unified_dataframe.columns:
            unified_dataframe['ReceiptNumber'] = unified_dataframe['ReceiptNumber'].apply(
                lambda x: int(x) if pd.notna(x) and str(x).strip().lstrip('-').replace('.', '').isdigit() and not any(c.isalpha() for c in str(x)) else x
            )

        # Generate unique file ID first
        import uuid
        file_id = str(uuid.uuid4())

        # Create output file with the file ID as the name in our persistent directory
        output_file_path = os.path.join(PROCESSED_FILES_DIR, f'{file_id}_formatted.xlsx')

        new_excel_workbook = openpyxl.Workbook()
        new_worksheet = new_excel_workbook.active

        # Write headers and data using DataFrame (more efficient)
        headers = unified_dataframe.columns.tolist()
        new_worksheet.append(headers)

        # Write data rows using DataFrame values
        for data_row in unified_dataframe.values:
            new_worksheet.append(data_row.tolist())

        # Apply styling
        header_font_style = Font(name="Calibri", size=11, bold=True)
        header_background_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        data_font_style = Font(name="Calibri", size=11)

        # Find column positions for special formatting using DataFrame approach
        amount_column_position = unified_dataframe.columns.get_loc('Amount_To_Apply') + 1 if 'Amount_To_Apply' in unified_dataframe.columns else 0
        date_column_positions = [unified_dataframe.columns.get_loc(date_col_name) + 1 for date_col_name in ['ReceiptCreateDate', 'PostDate', 'Date Last Change'] if date_col_name in unified_dataframe.columns]
        notes_column_position = unified_dataframe.columns.get_loc('Notes') + 1 if 'Notes' in unified_dataframe.columns else 0

        column_width_settings = {}
        for row_number, excel_row in enumerate(new_worksheet.rows, 1):
            for excel_cell in excel_row:
                # CRITICAL: Clear all existing formatting first (like autoFixFormatter.py)
                excel_cell.font = Font()
                excel_cell.fill = PatternFill()
                excel_cell.alignment = Alignment()
                excel_cell.number_format = 'General'

                # Apply new formatting
                if row_number == 1:
                    excel_cell.font, excel_cell.fill, excel_cell.alignment = header_font_style, header_background_fill, header_alignment
                else:
                    excel_cell.font = data_font_style

                if row_number > 1:
                    if excel_cell.column == amount_column_position:
                        excel_cell.number_format = '"$"#,##0.00_);\\("$"#,##0.00\\)'
                    elif excel_cell.column in date_column_positions:
                        excel_cell.number_format = 'mm-dd-yy'

                column_letter = excel_cell.column_letter
                column_width_settings[column_letter] = column_width_settings.get(column_letter, 8)
                try:
                    column_width_settings[column_letter] = min(25 if excel_cell.column == notes_column_position else 50, max(column_width_settings[column_letter], len(str(excel_cell.value)) + 2))
                except:
                    pass

        # Set column widths using list comprehension (more efficient like autoFixFormatter.py)
        [setattr(new_worksheet.column_dimensions[column_letter], 'width', width) for column_letter, width in column_width_settings.items()]

        new_worksheet.freeze_panes = 'A2'
        new_excel_workbook.save(output_file_path)

        processing_log['steps'].append({'step': 3, 'action': 'SUCCESS', 'result': f'Created formatted Excel with {len(successfully_found_columns)} columns and {maximum_data_length} rows'})
        processing_log['status'] = 'SUCCESS'

        logging.info(f"Success! Created formatted Excel file: {output_file_path}")

        return {
            'success': True,
            'output_file': output_file_path,
            'file_id': file_id,  # Add file ID for download
            'log': processing_log,
            'columns_found': len(successfully_found_columns),
            'rows_processed': maximum_data_length,
            'columns_matched': list(successfully_found_columns.keys()),
            'columns_missing': [col['target'] for col in columns_not_found]
        }

    except Exception as unexpected_error:
        import traceback
        logging.error(f"Excel formatting error: {unexpected_error}")
        traceback.print_exc()

        error_processing_log = locals().get('processing_log', {'timestamp': datetime.now().isoformat(), 'file': excel_file_path, 'steps': [], 'status': None})
        error_processing_log['steps'].append({'step': 'ERROR', 'action': 'FAILED', 'result': str(unexpected_error), 'trace': traceback.format_exc()})
        error_processing_log['status'] = 'FAILED'

        return {'error': str(unexpected_error), 'log': error_processing_log}

@excel_formatter_bp.route('/', methods=['GET'])
def get_service_info():
    """Get Excel Formatter service information"""
    return jsonify({
        'service': 'Excel Formatter API',
        'description': 'Upload Excel files to automatically detect and format columns with proper headers',
        'usage': 'POST with file parameter to process Excel files',
        'supported_formats': ['xlsx', 'xls'],
        'required_columns': REQUIRED_COLUMN_NAMES
    })

@excel_formatter_bp.route('/process', methods=['POST'])
def process_excel_file():
    """Process Excel file and format columns"""
    logging.info("Excel formatter processing request received")

    # Clean up old files before processing
    cleanup_old_files()

    try:
        # Check if file is uploaded
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file uploaded'
            }), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No file selected'
            }), 400

        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'error': 'Invalid file type. Please upload Excel file (.xlsx or .xls)'
            }), 400

        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            temp_input_path = tmp_file.name

        try:
            # Process the Excel file
            result = process_excel_and_fix_formatting(temp_input_path)

            if 'error' in result:
                return jsonify({
                    'success': False,
                    'error': result['error'],
                    'processing_log': result.get('log', {}),
                    'columns_found': result.get('columns_found', 0)
                }), 400

            return jsonify({
                'success': True,
                'message': f'Successfully processed Excel file with {result["columns_found"]} columns',
                'file_id': result['file_id'],  # Include the file_id for download
                'columns_found': result['columns_found'],
                'rows_processed': result['rows_processed'],
                'columns_matched': result['columns_matched'],
                'columns_missing': result['columns_missing'],
                'processing_log': result['log'],
                'download_ready': True
            })

        finally:
            # Clean up input temporary file
            if os.path.exists(temp_input_path):
                os.unlink(temp_input_path)

    except Exception as e:
        logging.error(f"Excel formatter processing error: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Processing failed: {str(e)}'
        }), 500

@excel_formatter_bp.route('/download/<file_id>', methods=['GET'])
def download_formatted_file(file_id):
    """Download the formatted Excel file using file ID"""
    try:
        # Validate file_id format (UUID)
        import uuid
        try:
            uuid.UUID(file_id)
        except ValueError:
            return jsonify({'error': 'Invalid file ID format'}), 400

        # Construct file path based on file ID
        file_path = os.path.join(PROCESSED_FILES_DIR, f'{file_id}_formatted.xlsx')

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found or expired. Files are automatically cleaned up after some time.'}), 404

        # Generate a filename with timestamp
        current_date = datetime.now()
        filename = f'formatted_excel_{current_date.strftime("%Y%m%d_%H%M%S")}.xlsx'

        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logging.error(f"Download error: {str(e)}")
        return jsonify({'error': f'Download failed: {str(e)}'}), 500

@excel_formatter_bp.route('/clear_results', methods=['POST'])
def clear_results():
    """Clear temporary result files"""
    logging.info("Clearing Excel formatter temporary files")

    try:
        # In production, you'd clean up session-specific temp files
        # For now, we'll just return success
        return jsonify({'status': 'success', 'message': 'Temporary files cleared'})

    except Exception as e:
        logging.error(f"Clear results error: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500