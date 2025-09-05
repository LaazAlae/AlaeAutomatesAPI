#!/usr/bin/env python3
"""
Production Statement Processing System

Optimized for O(n) performance and Railway deployment.
Memory-efficient, secure, and production-ready.

Version: 3.0 - Production Ready
"""

import fitz
import json
import re
import os
import sys
import gc
import logging
from datetime import datetime
from pathlib import Path
from difflib import get_close_matches, SequenceMatcher
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Set, Any
from collections import OrderedDict


class StatementProcessor:
    """
    Professional statement processor with O(n) complexity optimizations.
    
    Handles PDF text extraction, company matching against DNM lists,
    interactive questioning, and PDF splitting operations.
    """
    
    # Class constants for better performance and maintainability - FIXED TO MATCH MINIMAL
    US_STATES = frozenset([
        "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID",
        "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS",
        "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK",
        "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV",
        "WI", "WY", "DC"
    ])
    
    START_MARKERS = ["914.949.9618", "302.703.8961", "www.unitedcorporate.com", "AR@UNITEDCORPORATE.COM"]
    END_MARKER = "STATEMENT OF OPEN INVOICE(S)"
    # CRITICAL FIX: Added missing SKIP_LINES entries exactly like minimal version
    SKIP_LINES = {"Statement Date:", "Total Due:", "www.unitedcorporate.com", "Amount", "Invoice Number", "Description", "Invoice Date", "Invoice Number Description Invoice Date Amount"}
    
    def __init__(self, pdf_path: str, excel_path: str):
        """Initialize processor with file paths and pre-compile patterns for O(n) performance."""
        self.pdf_path = Path(pdf_path)
        self.excel_path = Path(excel_path)
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # Validate file paths immediately
        if not self.pdf_path.exists():
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        # Pre-compile regex patterns for optimal performance
        self._compile_patterns()
        
        # Load and pre-process DNM companies for O(1) lookups
        self.dnm_companies, self.normalized_company_map = self._load_dnm_companies()
        
        # Cache for processed pages to avoid reprocessing
        self._processed_pages: Set[int] = set()
        
        
        # Memory optimization: Clear unused references
        gc.collect()
    
    def _compile_patterns(self) -> None:
        """Pre-compile all regex patterns for maximum performance."""
        # Use the exact same business suffix list and pattern logic as the original code
        suffixes = [
            # Corporations
            'inc', 'incorporated', 'incorporation', 'corp', 'corporation',
            
            # Limited Liability Companies
            'llc', r'l\.l\.c\.?', 'limited liability company',
            
            # Limited Companies
            'ltd', 'limited', 'ltda',
            
            # Partnerships
            'llp', r'l\.l\.p\.?', 'limited liability partnership',
            'lp', r'l\.p\.?', 'limited partnership',
            'gp', 'general partnership',
            
            # Professional entities
            'pc', r'p\.c\.?', 'professional corporation',
            'pa', r'p\.a\.?', 'professional association',
            'pllc', r'p\.l\.l\.c\.?', 'professional limited liability company',
            'plc', r'p\.l\.c\.?', 'public limited company',
            'professional company',
            
            # General business
            'co', 'company', 'companies',
            'enterprise', 'enterprises',
            'group', 'groups',
            'holding', 'holdings',
            'international', 'intl',
            'global',
            'worldwide',
            'solutions',
            'services',
            'systems',
            'technologies', 'tech',
            'industries',
            
            # Specific entity types
            'sc', r's\.c\.?', 'service corporation',
            'bc', r'b\.c\.?', 'benefit corporation',
            'pbc', 'public benefit corporation',
            'nonprofit', 'non-profit',
            'foundation',
            'trust',
            'association', 'assn',
            'society',
            'institute',
            'academy',
            'center', 'centre',
            'organization', 'org',
            
            # Regional variations
            'pty', 'proprietary',
            'pvt', 'private',
            'pub', 'public',
            'joint venture', 'jv',
            'partnership',
            'syndicate',
            'consortium',
            'cooperative', 'coop', 'co-op',
            
            # Financial
            'bank', 'banking',
            'credit union',
            'mutual',
            'insurance', 'ins',
            'realty', 'real estate',
            'investment', 'investments',
            'capital',
            'financial', 'finance',
            
            # Other common endings
            'the', # Often appears at beginning or end
            'and', '&',
            'of',
            'dba', 'd/b/a', 'doing business as',
            'aka', 'a/k/a', 'also known as',
            'fka', 'f/k/a', 'formerly known as',
            'nka', 'n/k/a', 'now known as'
        ]
        
        # Use exact same logic as original code for pattern creation
        escaped_suffixes = []
        for suffix in suffixes:
            if suffix.startswith('r\'') or '\\' in suffix:
                # Already a raw string or contains escapes, use as-is
                escaped_suffixes.append(suffix)
            else:
                # Escape special regex characters
                escaped_suffixes.append(re.escape(suffix))
        
        # Create pattern that matches these suffixes at word boundaries
        pattern = r'\b(?:' + '|'.join(escaped_suffixes) + r')\b'
        
        self.patterns = {
            'page': re.compile(r'Page\s*(\d+)\s*of\s*(\d+)', re.IGNORECASE),
            'total_due': re.compile(r'(.+?)\s+Total Due\s+\$', re.IGNORECASE),
            'business_suffixes': re.compile(pattern, re.IGNORECASE),
            'clean_text': re.compile(r'[\s,.()\-_&]+'),
            'whitespace': re.compile(r'\s+')
        }
        
        # CRITICAL FIX: Enhanced patterns for company extraction exactly like minimal version
        self.PATTERNS = {
            'page': re.compile(r'Page\s*(\d+)\s*of\s*(\d+)', re.IGNORECASE),
            'total_due_subtotal': re.compile(r'Subtotal\s+\$[\d,]+\.\d{2}\s+([^\n\r]+?)\s+Total Due\s+\$[\d,]+\.\d{2}', re.IGNORECASE | re.MULTILINE),
            'total_due_multiline': re.compile(r'([^\n\r]+\n[^\n\r]*?)\s+Total Due\s+\$[\d,]+\.\d{2}', re.IGNORECASE | re.MULTILINE),
            'total_due_line': re.compile(r'(\S[^\n\r]*?)\s+Total Due\s+\$[\d,]+\.\d{2}', re.IGNORECASE | re.MULTILINE),
            'business_suffix': re.compile(r'\b(?:inc|incorporated|corp|corporation|llc|ltd|limited|llp|lp|pc|pa|pllc|plc|co|company|companies|enterprise|enterprises|group|groups|holding|holdings|international|intl|global|solutions|services|systems|technologies|tech|industries|foundation|trust|association|society|institute|center|centre|organization|org)\b', re.IGNORECASE),
            'clean_text': re.compile(r'[\s,.()\-_&]+'),
            'whitespace': re.compile(r'\s+')
        }
    
    def _normalize_company_name(self, name: str) -> str:
        """Normalize company names for consistent matching - O(1) operation."""
        if not name:
            return ""
        
        normalized = str(name).lower().strip()
        normalized = self.PATTERNS['business_suffix'].sub('', normalized)
        normalized = self.PATTERNS['clean_text'].sub('', normalized)
        
        return normalized.strip()
    
    def _load_dnm_companies(self) -> Tuple[List[str], Dict[str, str]]:
        """Load and pre-process DNM companies for O(1) lookups."""
        try:
            # Load Excel file with openpyxl instead of pandas
            workbook = load_workbook(self.excel_path, read_only=True)
            worksheet = workbook['10-2018']
            
            companies = []
            # Skip first 3 rows (2 header rows + 1 for 0-indexing)
            for row in worksheet.iter_rows(min_row=4, max_col=1, values_only=True):
                cell_value = row[0]
                if cell_value and str(cell_value).strip() and not str(cell_value).lower().startswith('name'):
                    companies.append(str(cell_value).strip())
            
            workbook.close()
            
            # Create normalized mapping for O(1) lookups
            normalized_map = {}
            for company in companies:
                normalized = self._normalize_company_name(company)
                if normalized:
                    normalized_map[normalized] = company
            
            return companies, normalized_map
            
        except Exception as e:
            raise RuntimeError(f"Failed to load DNM companies: {e}")
    
    def _find_company_match(self, company_name: str) -> Tuple[Optional[str], List[Dict[str, str]]]:
        """Find company match with O(1) exact match and enhanced fuzzy matching like minimal version."""
        # O(1) exact match check
        if company_name in self.dnm_companies:
            return company_name, []
        
        # O(1) normalized exact match
        normalized = self._normalize_company_name(company_name)
        if normalized in self.normalized_company_map:
            return self.normalized_company_map[normalized], []
        
        # Enhanced fuzzy matching: Check ALL companies above 60% threshold
        similar_matches = []
        if normalized:
            for norm_company, original_company in self.normalized_company_map.items():
                similarity_score = SequenceMatcher(None, normalized, norm_company).ratio() * 100
                if similarity_score >= 60.0:
                    similar_matches.append({
                        "company_name": original_company,
                        "percentage": f"{round(similarity_score, 1)}%"
                    })
            
            # Sort by similarity score (highest first)
            similar_matches.sort(key=lambda x: float(x["percentage"].replace('%', '')), reverse=True)
        
        return None, similar_matches
    
    def _detect_location(self, text: str) -> str:
        """Detect location using optimized state detection - O(k) where k = 50 states."""
        text_upper = f" {text.upper()} "
        return "National" if any(f" {state} " in text_upper for state in self.US_STATES) else "Foreign"
    
    def _determine_destination(self, exact_match: Optional[str], text: str, location: str, 
                            pages: int, best_percentage: float, has_email: bool) -> str:
        """Determine destination using business logic - O(1) operation."""
        if exact_match or has_email or best_percentage >= 90.0:
            return "DNM"
        elif location == "Foreign":
            return "Foreign"
        else:
            return "Natio Single" if pages == 1 else "Natio Multi"
    
    def _process_statement(self, text: str, page_num: int) -> Optional[Dict[str, Any]]:
        """CRITICAL FIX: Process statement exactly like minimal version."""
        page_match = self.PATTERNS['page'].search(text)
        current_page, total_pages = (int(page_match.group(1)), int(page_match.group(2))) if page_match else (1, 1)
        
        start_pos = min((text.find(marker) for marker in self.START_MARKERS if marker in text), default=-1)
        end_pos = text.find(self.END_MARKER)
        if start_pos == -1 or end_pos == -1:
            return None
        
        content = text[start_pos:end_pos]
        for marker in self.START_MARKERS:
            content = content.replace(marker, '')
        
        lines = [line.strip() for line in content.splitlines() if line.strip() and not any(skip in line for skip in self.SKIP_LINES)]
        if not lines:
            return None
        
        # Inline company extraction exactly like minimal
        fallback_company = lines[0].strip()
        extraction_method, fallback_used, fallback_reason = "unknown", False, ""
        
        match = self.PATTERNS['total_due_subtotal'].search(text)
        if match:
            company = self.PATTERNS['whitespace'].sub(' ', match.group(1).strip()).strip()
            if company.startswith("Amount "):
                company = company[7:].strip()
            extraction_method = "subtotal_pattern"
        else:
            match = self.PATTERNS['total_due_multiline'].search(text)
            if match:
                company = self.PATTERNS['whitespace'].sub(' ', match.group(1).replace('\n', ' ').strip()).strip()
                if company.startswith("Amount "):
                    company = company[7:].strip()
                extraction_method = "multiline_pattern"
                
                if len(company) > 100:
                    match = self.PATTERNS['total_due_line'].search(text)
                    if match:
                        company = self.PATTERNS['whitespace'].sub(' ', match.group(1).strip()).strip()
                        extraction_method = "line_pattern"
                    if len(company) > 100:
                        company, extraction_method, fallback_used, fallback_reason = fallback_company, "fallback", True, "Pattern extracted text too long"
            else:
                match = self.PATTERNS['total_due_line'].search(text)
                if match:
                    company = self.PATTERNS['whitespace'].sub(' ', match.group(1).strip()).strip()
                    if company.startswith("Amount "):
                        company = company[7:].strip()
                    extraction_method = "line_pattern"
                    
                    if len(company) > 100:
                        company, extraction_method, fallback_used, fallback_reason = fallback_company, "fallback", True, "Line pattern too long"
                else:
                    company, extraction_method, fallback_used, fallback_reason = fallback_company, "fallback", True, "No patterns found"
        
        # Use the better extraction method result
        
        rest_text = "\n".join(lines[1:])
        location = self._detect_location(rest_text)
        exact_match, similar_matches = self._find_company_match(company)
        
        # Calculate page range
        if total_pages == 1:
            page_range, first_page = str(page_num), page_num
        else:
            start_page = page_num - (current_page - 1)
            page_range, first_page = "-".join(map(str, range(start_page, start_page + total_pages))), start_page
        
        # Determine processing flags exactly like minimal version
        has_email = "email" in rest_text.lower()
        best_match = similar_matches[0] if similar_matches else None
        best_percentage = float(best_match["percentage"].replace('%', '')) if best_match else 0
        is_high_confidence = best_percentage >= 90.0
        
        manual_required, ask_question = False, False
        if not (has_email or is_high_confidence or exact_match):
            manual_required = len(similar_matches) > 0
            if manual_required:
                ask_question = best_percentage < 90.0
        
        if exact_match or has_email or is_high_confidence:
            destination = "DNM"
        elif location == "Foreign":
            destination = "Foreign"
        else:
            destination = "Natio Single" if total_pages == 1 else "Natio Multi"
        
        # CRITICAL FIX: Use OrderedDict with exact field ordering like minimal
        result = OrderedDict()
        result["company_name"] = company
        if company.strip() != fallback_company.strip():
            result["unusedCompanyName"] = fallback_company
        result["exact_match"] = exact_match
        result["similar_matches"] = similar_matches
        result["manual_required"] = manual_required
        result["ask_question"] = ask_question
        result["rest_of_lines"] = rest_text
        result["location"] = location
        result["paging"] = f"page {current_page} of {total_pages}"
        result["number_of_pages"] = str(total_pages)
        result["page_number_in_uploaded_pdf"] = page_range
        result["first_page_number"] = first_page
        result["destination"] = destination
        result["extraction_method"] = extraction_method
        result["fallbackUsed"] = fallback_used
        if fallback_used:
            result["fallbackReason"] = fallback_reason
            
        return result
    
    def extract_statements(self) -> List[Dict[str, Any]]:
        """CRITICAL FIX: Extract statements exactly like minimal version - jump to last page."""
        doc = fitz.open(str(self.pdf_path))
        statements = []
        processed_pages = set()
        
        for page_idx in range(len(doc)):
            page_num = page_idx + 1
            if page_num in processed_pages:
                continue
            
            page_text = doc.load_page(page_idx).get_text()
            
            # Inline boundary detection
            page_match = self.PATTERNS['page'].search(page_text)
            if not page_match:
                continue
            
            start_pos = min((page_text.find(marker) for marker in self.START_MARKERS if marker in page_text), default=-1)
            end_pos = page_text.find(self.END_MARKER)
            if start_pos == -1 or end_pos == -1:
                continue
            
            total_pages = int(page_match.group(2))
            current_page = int(page_match.group(1))
            start_page = page_num - (current_page - 1)
            last_page_num = start_page + total_pages - 1
            
            # CRITICAL FIX: Jump to last page like minimal version
            if last_page_num <= len(doc):
                last_page_text = doc.load_page(last_page_num - 1).get_text()
                statement_data = self._process_statement(last_page_text, last_page_num)
                
                if statement_data:
                    statements.append(statement_data)
            
            processed_pages.update(range(start_page, last_page_num + 1))
        
        doc.close()
        
        
        return statements
    
    def process_interactive_questions(self, statements: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Process interactive questions for companies requiring manual review."""
        questions_needed = [stmt for stmt in statements if stmt.get('ask_question', False)]
        
        if not questions_needed:
            print("No manual questions required.")
            return statements
        
        print(f"\nFound {len(questions_needed)} companies requiring manual review:")
        
        skip_all = False
        history = []  # Track history of statement states
        i = 0  # Current question index
        
        while i < len(questions_needed):
            if skip_all:
                for j in range(i, len(questions_needed)):
                    questions_needed[j]['user_answered'] = 'skip'
                break
            
            statement = questions_needed[i]
            company_name = statement.get('company_name', 'Unknown')
            similar_matches = statement.get('similar_matches', [])
            
            if similar_matches:
                best_match = similar_matches[0]
                print(f"\nQuestion {i + 1} of {len(questions_needed)}:")
                print(f"Company '{company_name}' is similar to '{best_match['company_name']}' ({best_match['percentage']})")
                print("Are they the same company? (y/n/s to skip all/p to go back)")
                
                while True:
                    try:
                        response = input("> ").strip().lower()
                        
                        if response == 'y':
                            # Save current state to history before making changes
                            history.append({
                                'index': i,
                                'statement_state': {
                                    'destination': statement.get('destination'),
                                    'user_answered': statement.get('user_answered')
                                }
                            })
                            
                            statement['destination'] = 'DNM'
                            statement['user_answered'] = 'yes'
                            print(f" Marked '{company_name}' as DNM")
                            i += 1  # Move to next question
                            break
                            
                        elif response == 'n':
                            # Save current state to history before making changes
                            history.append({
                                'index': i,
                                'statement_state': {
                                    'destination': statement.get('destination'),
                                    'user_answered': statement.get('user_answered')
                                }
                            })
                            
                            statement['user_answered'] = 'no'
                            print(f" Kept '{company_name}' as {statement['destination']}")
                            i += 1  # Move to next question
                            break
                            
                        elif response == 's':
                            skip_all = True
                            statement['user_answered'] = 'skip'
                            print(" Skipping remaining questions")
                            break
                            
                        elif response == 'p':
                            if not history:
                                print("No previous questions to go back to")
                                continue
                            
                            # Restore previous state
                            previous = history.pop()
                            prev_index = previous['index']
                            prev_state = previous['statement_state']
                            
                            # Restore the previous statement's state
                            prev_statement = questions_needed[prev_index]
                            prev_statement['destination'] = prev_state['destination']
                            if 'user_answered' in prev_state and prev_state['user_answered'] is not None:
                                prev_statement['user_answered'] = prev_state['user_answered']
                            elif 'user_answered' in prev_statement:
                                del prev_statement['user_answered']
                            
                            i = prev_index  # Go back to previous question
                            print(f"↩ Going back to question {i + 1}")
                            break
                            
                        else:
                            print("Please enter 'y', 'n', 's', or 'p'")
                            
                    except (KeyboardInterrupt, EOFError):
                        print("\nOperation cancelled.")
                        sys.exit(0)
        
        return statements
    
    def create_split_pdfs(self, statements: List[Dict[str, Any]]) -> Dict[str, int]:
        """Split PDF into destination-based files - O(n) operation."""
        # Group statements by destination - O(n)
        destinations = {"DNM": [], "Foreign": [], "Natio Single": [], "Natio Multi": []}
        
        for statement in statements:
            dest = statement.get('destination', '').strip()
            if dest in destinations:
                destinations[dest].append(statement)
        
        # Create output PDFs
        output_files = {
            "DNM": "DNM.pdf",
            "Foreign": "Foreign.pdf", 
            "Natio Single": "natioSingle.pdf",
            "Natio Multi": "natioMulti.pdf"
        }
        
        results = {}
        
        try:
            reader = PdfReader(str(self.pdf_path))
            total_pages = len(reader.pages)
            
            for dest, statements_list in destinations.items():
                if not statements_list:
                    continue
                
                writer = PdfWriter()
                pages_added = 0
                
                for statement in statements_list:
                    page_range = statement.get('page_number_in_uploaded_pdf', '')
                    for page_str in page_range.split('-'):
                        try:
                            page_num = int(page_str.strip()) - 1  # Convert to 0-based index
                            if 0 <= page_num < total_pages:
                                writer.add_page(reader.pages[page_num])
                                pages_added += 1
                        except ValueError:
                            continue
                
                if pages_added > 0:
                    output_path = output_files[dest]
                    with open(output_path, 'wb') as f:
                        writer.write(f)
                    results[dest] = pages_added
                    print(f" Created {output_path} with {pages_added} pages")
            
            return results
            
        except Exception as e:
            raise RuntimeError(f"Failed to create split PDFs: {e}")
    
    def save_results(self, statements: List[Dict[str, Any]], output_path: Optional[str] = None) -> str:
        """Save processing results to JSON file."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = Path(f"output_{timestamp}")
        output_dir.mkdir(exist_ok=True)
        
        # Clean up internal logging data
        for statement in statements:
            if '_extraction_log' in statement:
                del statement['_extraction_log']
        
        # Save clean JSON for API consumers
        data = {
            "dnm_companies": self.dnm_companies,
            "extracted_statements": statements,
            "total_statements_found": len(statements),
            "processing_timestamp": datetime.now().isoformat()
        }
        
        json_path = output_dir / "results.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f" Results saved to {json_path}")
        return str(json_path)
    
    def run_complete_workflow(self, skip_questions: bool = False) -> bool:
        """Execute the complete statement processing workflow."""
        try:
            print("=" * 60)
            print("          PROFESSIONAL STATEMENT PROCESSOR")
            print("=" * 60)
            print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print()
            
            # Step 1: Extract statements
            print(" Step 1: Extracting statements from PDF...")
            statements = self.extract_statements()
            print(f" Extracted {len(statements)} statements")
            
            # Step 2: Process interactive questions (skip if requested)
            if not skip_questions:
                print("\n Step 2: Processing manual questions...")
                statements = self.process_interactive_questions(statements)
                print(" Manual questions processed")
            else:
                print("\n Step 2: Skipping interactive questions...")
                print(" Questions skipped for comparison")
            
            # Step 3: Save results
            print("\n Step 3: Saving results...")
            output_file = self.save_results(statements)
            print(" Results saved")
            
            if not skip_questions:
                # Step 4: Create split PDFs
                print("\n Step 4: Creating destination PDFs...")
                split_results = self.create_split_pdfs(statements)
                print(" PDFs created successfully")
                
                # Final summary
                print("\n" + "=" * 60)
                print(" WORKFLOW COMPLETED SUCCESSFULLY!")
                print("=" * 60)
                
                print(f"Total statements processed: {len(statements)}")
                print(f"JSON output: {output_file}")
                print("PDF outputs created:")
                for dest, pages in split_results.items():
                    print(f"  • {dest}: {pages} pages")
            else:
                # Comparison mode summary
                print("\n" + "=" * 60)
                print(" EXTRACTION COMPLETED FOR COMPARISON")
                print("=" * 60)
                
                manual_count = sum(1 for s in statements if s.get('manual_required', False))
                ask_count = sum(1 for s in statements if s.get('ask_question', False))
                
                print(f"Total statements processed: {len(statements)}")
                print(f"Manual review required: {manual_count}")
                print(f"Ask question required: {ask_count}")
                print(f"JSON output: {output_file}")
            
            print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
            return True
            
        except Exception as e:
            print(f"\n Workflow failed: {e}")
            return False


def find_files_in_directory() -> Tuple[Optional[str], Optional[str]]:
    """Find PDF and Excel files in current directory."""
    pdf_files = list(Path('.').glob('*.pdf'))
    excel_files = list(Path('.').glob('*.xlsx')) + list(Path('.').glob('*.xls'))
    
    if pdf_files and excel_files:
        return str(pdf_files[0]), str(excel_files[0])
    return None, None


def get_file_paths() -> Tuple[str, str]:
    """Get file paths from user input or auto-detect."""
    # Try auto-detection first
    pdf_path, excel_path = find_files_in_directory()
    
    if pdf_path and excel_path:
        print(f"Found files:")
        print(f"  PDF: {pdf_path}")
        print(f"  Excel: {excel_path}")
        
        use_files = input("Use these files? (y/n): ").strip().lower()
        if use_files == 'y':
            return pdf_path, excel_path
    
    # Manual file selection
    print("\nManual file selection:")
    
    while True:
        pdf_path = input("Enter PDF file path: ").strip().strip('"')
        if os.path.exists(pdf_path) and pdf_path.lower().endswith('.pdf'):
            break
        print(" Invalid PDF file path")
    
    while True:
        excel_path = input("Enter Excel file path: ").strip().strip('"')
        if os.path.exists(excel_path) and excel_path.lower().endswith(('.xlsx', '.xls')):
            break
        print(" Invalid Excel file path")
    
    return pdf_path, excel_path


def main() -> int:
    """Main entry point for the statement processor."""
    try:
        print("Professional Statement Processing System v2.0")
        print("=" * 50)
        
        # Get file paths
        pdf_path, excel_path = get_file_paths()
        
        # Create and run processor
        processor = StatementProcessor(pdf_path, excel_path)
        skip_questions = '--skip-questions' in sys.argv
        success = processor.run_complete_workflow(skip_questions)
        
        return 0 if success else 1
        
    except KeyboardInterrupt:
        print("\n\n Process interrupted by user")
        return 1
    except Exception as e:
        print(f"\n Fatal error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())