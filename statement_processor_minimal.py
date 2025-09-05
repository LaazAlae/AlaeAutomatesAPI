"""
MINIMAL STATEMENT PROCESSOR - Ultra-Simplified Architecture
============================================================

FILE STRUCTURE:
1. SETUP & PATTERNS - Global constants and regex patterns
2. DATA LOADING - load_excel_data()
3. PDF PROCESSING - extract_statements(), process_statement()
4. USER INTERACTION - ask_questions()
5. OUTPUT GENERATION - save_outputs()
6. MAIN WORKFLOW - main()

WORKFLOW SUMMARY:
=================
1. AUTO-DETECT FILES: Finds PDF and Excel files in current directory
2. LOAD EXCEL: Reads company list from Excel, creates O(1) lookup map for instant matching
3. SCAN PDF: Reads PDF pages, detects statement boundaries using page markers
4. JUMP TO LAST PAGE: For each statement, jumps directly to final page (most efficient)
5. EXTRACT COMPANY: Uses 4 regex patterns in priority order to extract company names
6. MATCH COMPANIES: Checks exact match first, then similarity matching (60%+ threshold)
7. CLASSIFY DESTINATION: Routes to DNM/Foreign/National based on matches and location
8. ASK QUESTIONS: For uncertain matches, asks user for confirmation
9. GENERATE OUTPUTS: Creates timestamped folder with JSON results and sorted PDFs

FUNCTION DETAILS:
=================
- load_excel_data(): Loads companies, creates normalized lookup map, normalizes names
- extract_statements(): Scans PDF, jumps to last page, processes each statement
- process_statement(): Extracts company, finds matches, determines destination, builds result
- ask_questions(): Interactive prompts for uncertain company matches
- save_outputs(): Creates timestamped folder, saves JSON and PDFs
- main(): Orchestrates entire workflow, handles errors and user interruption
"""











import fitz, json, re, os, sys
from datetime import datetime
from pathlib import Path
from difflib import SequenceMatcher
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Any
from collections import OrderedDict
################################################################################











# SETUP & PATTERNS
################################################################################

PATTERNS = {
    'page': re.compile(r'Page\s*(\d+)\s*of\s*(\d+)', re.IGNORECASE),
    'total_due_subtotal': re.compile(r'Subtotal\s+\$[\d,]+\.\d{2}\s+([^\n\r*]+?)\s+Total Due\s+\$[\d,]+\.\d{2}', re.IGNORECASE | re.MULTILINE),
    'total_due_multiline': re.compile(r'([^\n\r*]+\n[^\n\r*]*?)\s+Total Due\s+\$[\d,]+\.\d{2}', re.IGNORECASE | re.MULTILINE),
    'total_due_line': re.compile(r'(\S[^\n\r*]*?)\s+Total Due\s+\$[\d,]+\.\d{2}', re.IGNORECASE | re.MULTILINE),
    'business_suffix': re.compile(r'\b(?:inc|incorporated|corp|corporation|llc|ltd|limited|llp|lp|pc|pa|pllc|plc|co|company|companies|enterprise|enterprises|group|groups|holding|holdings|international|intl|global|solutions|services|systems|technologies|tech|industries|foundation|trust|association|society|institute|center|centre|organization|org)\b', re.IGNORECASE),
    'clean_text': re.compile(r'[\s,.()\-_&]+'),
    'whitespace': re.compile(r'\s+')
}

US_STATES = frozenset(["AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "PR", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY", "DC"])

START_MARKERS = ["914.949.9618", "302.703.8961", "www.unitedcorporate.com", "AR@UNITEDCORPORATE.COM"]
END_MARKER = "STATEMENT OF OPEN INVOICE(S)"
SKIP_LINES = {"Statement Date:", "Total Due:", "www.unitedcorporate.com", "Amount", "Invoice Number", "Description", "Invoice Date", "Invoice Number Description Invoice Date Amount"}

################################################################################










# DATA LOADING
################################################################################

def load_excel_data(excel_path: str) -> Tuple[List[str], Dict[str, str]]:
    workbook = load_workbook(excel_path, read_only=True)
    worksheet = workbook['10-2018']
    
    companies = []
    normalized_map = {}
    
    for row in worksheet.iter_rows(min_row=4, max_col=1, values_only=True):
        if row[0] and str(row[0]).strip():
            company = str(row[0]).strip()
            companies.append(company)
            
            # Normalize inline
            normalized = str(company).lower().strip()
            normalized = PATTERNS['business_suffix'].sub('', normalized)
            normalized = PATTERNS['clean_text'].sub('', normalized).strip()
            if normalized:
                normalized_map[normalized] = company
    
    workbook.close()
    return companies, normalized_map

################################################################################











# PDF PROCESSING
################################################################################

def extract_statements(pdf_path: str, dnm_companies: List[str], normalized_map: Dict[str, str]) -> Tuple[List[Dict[str, Any]], List[Dict]]:
    doc = fitz.open(pdf_path)
    statements = []
    processed_pages = set()
    extraction_log = []
    
    for page_idx in range(len(doc)):
        page_num = page_idx + 1
        if page_num in processed_pages:
            continue
        
        page_text = doc.load_page(page_idx).get_text()
        
        # Inline boundary detection
        page_match = PATTERNS['page'].search(page_text)
        if not page_match:
            continue
        
        start_pos = min((page_text.find(marker) for marker in START_MARKERS if marker in page_text), default=-1)
        end_pos = page_text.find(END_MARKER)
        if start_pos == -1 or end_pos == -1:
            continue
        
        total_pages = int(page_match.group(2))
        current_page = int(page_match.group(1))
        start_page = page_num - (current_page - 1)
        last_page_num = start_page + total_pages - 1
        
        if last_page_num <= len(doc):
            last_page_text = doc.load_page(last_page_num - 1).get_text()
            statement_data = process_statement(last_page_text, last_page_num, dnm_companies, normalized_map, extraction_log)
            
            if statement_data:
                statements.append(statement_data)
        
        processed_pages.update(range(start_page, last_page_num + 1))
    
    doc.close()
    
    # Return both statements and extraction log separately
    return statements, extraction_log

def process_statement(text: str, page_num: int, dnm_companies: List[str], normalized_map: Dict[str, str], extraction_log: List) -> Optional[Dict[str, Any]]:
    page_match = PATTERNS['page'].search(text)
    current_page, total_pages = (int(page_match.group(1)), int(page_match.group(2))) if page_match else (1, 1)
    
    start_pos = min((text.find(marker) for marker in START_MARKERS if marker in text), default=-1)
    end_pos = text.find(END_MARKER)
    if start_pos == -1 or end_pos == -1:
        return None
    
    content = text[start_pos:end_pos]
    for marker in START_MARKERS:
        content = content.replace(marker, '')
    
    lines = [line.strip() for line in content.splitlines() if line.strip() and not any(skip in line for skip in SKIP_LINES)]
    if not lines:
        return None
    
    # Inline company extraction
    fallback_company = lines[0].strip()
    extraction_method, fallback_used, fallback_reason = "unknown", False, ""
    
    match = PATTERNS['total_due_subtotal'].search(text)
    if match:
        company = PATTERNS['whitespace'].sub(' ', match.group(1).strip()).strip()
        if company.startswith("Amount "):
            company = company[7:].strip()
        extraction_method = "subtotal_pattern"
    else:
        match = PATTERNS['total_due_multiline'].search(text)
        if match:
            company = PATTERNS['whitespace'].sub(' ', match.group(1).replace('\n', ' ').strip()).strip()
            if company.startswith("Amount "):
                company = company[7:].strip()
            extraction_method = "multiline_pattern"
            
            if len(company) > 100:
                match = PATTERNS['total_due_line'].search(text)
                if match:
                    company = PATTERNS['whitespace'].sub(' ', match.group(1).strip()).strip()
                    extraction_method = "line_pattern"
                if len(company) > 100:
                    company, extraction_method, fallback_used, fallback_reason = fallback_company, "fallback", True, "Pattern extracted text too long"
        else:
            match = PATTERNS['total_due_line'].search(text)
            if match:
                company = PATTERNS['whitespace'].sub(' ', match.group(1).strip()).strip()
                if company.startswith("Amount "):
                    company = company[7:].strip()
                extraction_method = "line_pattern"
                
                if len(company) > 100:
                    company, extraction_method, fallback_used, fallback_reason = fallback_company, "fallback", True, "Line pattern too long"
            else:
                company, extraction_method, fallback_used, fallback_reason = fallback_company, "fallback", True, "No patterns found"
    
    if company.strip() != fallback_company.strip():
        extraction_log.append({
            'page_num': page_num, 'current_page': current_page, 'total_pages': total_pages,
            'old_method': fallback_company, 'new_method': company,
            'extraction_method': extraction_method, 'match': company.strip() == fallback_company.strip()
        })
    
    rest_text = "\n".join(lines[1:])
    
    # Inline location detection
    text_upper = f" {rest_text.upper()} "
    location = "National" if any(f" {state} " in text_upper for state in US_STATES) else "Foreign"
    
    # Inline matching
    if company in dnm_companies:
        exact_match, similar_matches = company, []
    else:
        normalized = str(company).lower().strip()
        normalized = PATTERNS['business_suffix'].sub('', normalized)
        normalized = PATTERNS['clean_text'].sub('', normalized).strip()
        
        if normalized in normalized_map:
            exact_match, similar_matches = normalized_map[normalized], []
        else:
            exact_match, similar_matches = None, []
            if normalized:
                for norm_company, original_company in normalized_map.items():
                    similarity_score = SequenceMatcher(None, normalized, norm_company).ratio() * 100
                    if similarity_score >= 60.0:
                        similar_matches.append({"company_name": original_company, "percentage": f"{round(similarity_score, 1)}%"})
                
                similar_matches.sort(key=lambda x: float(x["percentage"].replace('%', '')), reverse=True)
    
    if total_pages == 1:
        page_range, first_page = str(page_num), page_num
    else:
        start_page = page_num - (current_page - 1)
        page_range, first_page = "-".join(map(str, range(start_page, start_page + total_pages))), start_page
    
    has_email = "email" in rest_text.lower()
    best_match = similar_matches[0] if similar_matches else None
    best_percentage = float(best_match["percentage"].replace('%', '')) if best_match else 0
    is_high_confidence = best_percentage >= 90.0
    
    manual_required, ask_question = False, False
    if not (has_email or is_high_confidence or exact_match):
        manual_required = len(similar_matches) > 0
        if manual_required:
            ask_question = best_percentage < 90.0
    
    if exact_match or has_email or is_high_confidence:
        destination = "DNM"
    elif location == "Foreign":
        destination = "Foreign"
    else:
        destination = "Natio Single" if total_pages == 1 else "Natio Multi"
    
    result = OrderedDict()
    result["company_name"] = company
    if company.strip() != fallback_company.strip():
        result["unusedCompanyName"] = fallback_company
    result["exact_match"] = exact_match
    result["similar_matches"] = similar_matches
    result["manual_required"] = manual_required
    result["ask_question"] = ask_question
    result["rest_of_lines"] = rest_text
    result["location"] = location
    result["paging"] = f"page {current_page} of {total_pages}"
    result["number_of_pages"] = str(total_pages)
    result["page_number_in_uploaded_pdf"] = page_range
    result["first_page_number"] = first_page
    result["destination"] = destination
    result["extraction_method"] = extraction_method
    result["fallbackUsed"] = fallback_used
    if fallback_used:
        result["fallbackReason"] = fallback_reason
        
    return result

################################################################################












################################################################################











# USER INTERACTION - Interactive question handling with back navigation and history tracking
################################################################################

def ask_questions(statements: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    questions_needed = [stmt for stmt in statements if stmt.get('ask_question', False)]
    if not questions_needed:
        print("No manual questions required.")
        return statements
    
    print(f"\nFound {len(questions_needed)} companies requiring manual review:")
    
    skip_all = False
    history = []  # Track history of statement states
    i = 0  # Current question index
    
    while i < len(questions_needed):
        if skip_all:
            for j in range(i, len(questions_needed)):
                questions_needed[j]['user_answered'] = 'skip'
            break
        
        statement = questions_needed[i]
        company_name = statement.get('company_name', 'Unknown')
        similar_matches = statement.get('similar_matches', [])
        
        if similar_matches:
            best_match = similar_matches[0]
            print(f"\nQuestion {i + 1} of {len(questions_needed)}:")
            print(f"Company '{company_name}' is similar to '{best_match['company_name']}' ({best_match['percentage']})")
            print("Are they the same company? (y/n/s to skip all/p to go back)")
            
            while True:
                try:
                    response = input("> ").strip().lower()
                    
                    if response == 'y':
                        # Save current state to history before making changes
                        history.append({
                            'index': i,
                            'statement_state': {
                                'destination': statement.get('destination'),
                                'user_answered': statement.get('user_answered')
                            }
                        })
                        
                        statement['destination'] = 'DNM'
                        statement['user_answered'] = 'yes'
                        print(f" Marked '{company_name}' as DNM")
                        i += 1  # Move to next question
                        break
                        
                    elif response == 'n':
                        # Save current state to history before making changes
                        history.append({
                            'index': i,
                            'statement_state': {
                                'destination': statement.get('destination'),
                                'user_answered': statement.get('user_answered')
                            }
                        })
                        
                        statement['user_answered'] = 'no'
                        print(f" Kept '{company_name}' as {statement['destination']}")
                        i += 1  # Move to next question
                        break
                        
                    elif response == 's':
                        skip_all = True
                        statement['user_answered'] = 'skip'
                        print(" Skipping remaining questions")
                        break
                        
                    elif response == 'p':
                        if not history:
                            print("No previous questions to go back to")
                            continue
                        
                        # Restore previous state
                        previous = history.pop()
                        prev_index = previous['index']
                        prev_state = previous['statement_state']
                        
                        # Restore the previous statement's state
                        prev_statement = questions_needed[prev_index]
                        prev_statement['destination'] = prev_state['destination']
                        if 'user_answered' in prev_state and prev_state['user_answered'] is not None:
                            prev_statement['user_answered'] = prev_state['user_answered']
                        elif 'user_answered' in prev_statement:
                            del prev_statement['user_answered']
                        
                        i = prev_index  # Go back to previous question
                        print(f"↩ Going back to question {i + 1}")
                        break
                        
                    else:
                        print("Please enter 'y', 'n', 's', or 'p'")
                        
                except (KeyboardInterrupt, EOFError):
                    print("\nOperation cancelled.")
                    sys.exit(0)
    
    return statements

################################################################################











# OUTPUT GENERATION - Creates timestamped folders with JSON and PDF outputs
################################################################################

def save_outputs(pdf_path: str, statements: List[Dict[str, Any]], dnm_companies: List[str], extraction_log: List = None, skip_pdfs: bool = False) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = Path(f"output_{timestamp}")
    output_dir.mkdir(exist_ok=True)
    
    # Use passed extraction_log or empty list
    if extraction_log is None:
        extraction_log = []
    
    # Save JSON
    data = {
        "dnm_companies": dnm_companies,
        "extracted_statements": statements,
        "total_statements_found": len(statements),
        "processing_timestamp": datetime.now().isoformat(),
        "extraction_comparison_log": {
            "total_statements_with_different_extractions": len(extraction_log),
            "extraction_details": extraction_log,
            "summary": {
                "extraction_methods_used": list(set(comp['extraction_method'] for comp in extraction_log)),
                "pages_with_multiline_extraction": len([c for c in extraction_log if c['extraction_method'] == 'multiline_pattern']),
                "pages_with_subtotal_extraction": len([c for c in extraction_log if c['extraction_method'] == 'subtotal_pattern']),
                "pages_with_improved_accuracy": len([c for c in extraction_log if c['extraction_method'] != 'fallback'])
            }
        }
    }
    
    json_path = output_dir / "results.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    
    if skip_pdfs:
        return f"Completed: {len(statements)} statements, JSON: {json_path}"
    
    # Create PDFs
    destinations = {"DNM": [], "Foreign": [], "Natio Single": [], "Natio Multi": []}
    for statement in statements:
        dest = statement.get('destination', '').strip()
        if dest in destinations:
            destinations[dest].append(statement)
    
    output_files = {"DNM": "DNM.pdf", "Foreign": "Foreign.pdf", "Natio Single": "NatioSingle.pdf", "Natio Multi": "NatioMulti.pdf"}
    results = {}
    reader = PdfReader(pdf_path)
    total_pages = len(reader.pages)
    
    for dest, statements_list in destinations.items():
        if not statements_list:
            continue
        
        writer = PdfWriter()
        pages_added = 0
        
        for statement in statements_list:
            page_range = statement.get('page_number_in_uploaded_pdf', '')
            for page_str in page_range.split('-'):
                try:
                    page_num = int(page_str.strip()) - 1
                    if 0 <= page_num < total_pages:
                        writer.add_page(reader.pages[page_num])
                        pages_added += 1
                except ValueError:
                    continue
        
        if pages_added > 0:
            output_path = output_dir / output_files[dest]
            with open(output_path, 'wb') as f:
                writer.write(f)
            results[dest] = pages_added
    
    return f"Completed: {len(statements)} statements, JSON: {json_path}\nPDFs: {sum(results.values())} total pages in {output_dir}"

################################################################################











# MAIN WORKFLOW - Complete processing pipeline with professional output
################################################################################

def main() -> int:
    try:
        print("=" * 60)
        print("          MINIMAL STATEMENT PROCESSOR")
        print("=" * 60)
        print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()
        
        # Auto-detect files
        pdf_files = list(Path('.').glob('*.pdf'))
        excel_files = list(Path('.').glob('*.xlsx')) + list(Path('.').glob('*.xls'))
        
        if not (pdf_files and excel_files):
            pdf_path = os.environ.get('PDF_PATH')
            excel_path = os.environ.get('EXCEL_PATH')
            if not (pdf_path and excel_path):
                raise FileNotFoundError("No PDF or Excel files found")
        else:
            pdf_path, excel_path = str(pdf_files[0]), str(excel_files[0])
        
        print(f"Files found: {pdf_path}, {excel_path}")
        print()
        
        # Step 1: Load Excel data
        print(" Step 1: Loading Excel data...")
        dnm_companies, normalized_map = load_excel_data(excel_path)
        print(f" Loaded {len(dnm_companies)} DNM companies")
        
        # Step 2: Extract statements
        print("\n Step 2: Extracting statements from PDF...")
        statements, extraction_log = extract_statements(pdf_path, dnm_companies, normalized_map)
        print(f" Extracted {len(statements)} statements")
        
        # Step 3: Process interactive questions (skip if requested)
        skip_questions = '--skip-questions' in sys.argv
        if not skip_questions:
            print("\n Step 3: Processing manual questions...")
            statements = ask_questions(statements)
            print(" Manual questions processed")
        else:
            print("\n Step 3: Skipping interactive questions...")
            print(" Questions skipped for comparison")
        
        # Step 4: Save outputs
        print("\n Step 4: Saving results...")
        result = save_outputs(pdf_path, statements, dnm_companies, extraction_log, skip_questions)
        print(" Results saved")
        
        # Final summary
        if not skip_questions:
            print("\n" + "=" * 60)
            print(" WORKFLOW COMPLETED SUCCESSFULLY!")
            print("=" * 60)
            print(f"Total statements processed: {len(statements)}")
        else:
            print("\n" + "=" * 60)
            print(" EXTRACTION COMPLETED FOR COMPARISON")
            print("=" * 60)
            
            manual_count = sum(1 for s in statements if s.get('manual_required', False))
            ask_count = sum(1 for s in statements if s.get('ask_question', False))
            
            print(f"Total statements processed: {len(statements)}")
            print(f"Manual review required: {manual_count}")
            print(f"Ask question required: {ask_count}")
        
        print(f"\n{result}")
        print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        return 0
        
    except KeyboardInterrupt:
        print("\nCancelled")
        return 1
    except Exception as e:
        print(f"Error: {e}")
        return 1
    
################################################################################










if __name__ == "__main__":
    sys.exit(main())