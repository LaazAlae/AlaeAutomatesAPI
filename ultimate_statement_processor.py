#!/usr/bin/env python3
"""
ULTIMATE Statement Processing System - Production Ready

Combines the best of both worlds:
- Accurate company name extraction from data.py (with multiline & subtotal patterns)  
- Comprehensive company matching from statement_processor.py (ALL companies above 60%)
- Production-ready features with O(1) optimizations
- Complete extraction comparison logging integrated into JSON output

Version: ULTIMATE - Ready for Production Deployment
"""

import fitz
import json
import re
import os
import sys
import gc
import logging
from datetime import datetime
from pathlib import Path
from difflib import get_close_matches, SequenceMatcher
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Set, Any
from collections import OrderedDict


################################################################################












# ULTIMATE STATEMENT PROCESSOR CLASS
################################################################################

class StatementProcessor:
    """
    Ultimate statement processor combining accurate extraction with comprehensive matching.
    
    Features:
    - Fixed company name extraction (multiline + subtotal patterns)
    - ALL company matching above 60% threshold
    - Integrated extraction comparison logging
    - O(1) optimizations throughout
    - Production-ready error handling
    """
    
################################################################################













#### INITIALIZATION AND SETUP
################################################################################
    
    def __init__(self, pdf_path: str, excel_path: str):
        """Initialize with optimized patterns and pre-processed company data."""
        self.pdf_path = Path(pdf_path)
        self.excel_path = Path(excel_path)
        
        # Validate paths immediately
        if not self.pdf_path.exists():
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
#### Regex Pattern Setup (ULTIMATE - Fixed Extraction Patterns)
        self.patterns = {
            'page': re.compile(r'Page\s*(\d+)\s*of\s*(\d+)', re.IGNORECASE),
            'total_due_from_subtotal': re.compile(r'Subtotal\s+\$[\d,]+\.\d{2}\s+([^\n\r]+?)\s+Total Due\s+\$[\d,]+\.\d{2}', re.IGNORECASE | re.MULTILINE),
            'total_due_multiline': re.compile(r'([^\n\r]+\n[^\n\r]*?)\s+Total Due\s+\$[\d,]+\.\d{2}', re.IGNORECASE | re.MULTILINE),
            'total_due_line_end': re.compile(r'(\S[^\n\r]*?)\s+Total Due\s+\$[\d,]+\.\d{2}', re.IGNORECASE | re.MULTILINE),
            'total_due_fallback': re.compile(r'(.+?)\s+Total Due\s+\$', re.IGNORECASE),
            'business_suffixes': self._create_business_suffix_pattern(),
            'clean_text': re.compile(r'[\s,.()\-_&]+'),
            'whitespace': re.compile(r'\s+')
        }

#### US States and Text Markers (O(1) Lookups)
        self.us_states = frozenset([
            "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID",
            "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", 
            "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK",
            "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV",
            "WI", "WY", "DC"
        ])
        
        self.start_markers = ["914.949.9618", "302.703.8961", "www.unitedcorporate.com", "AR@UNITEDCORPORATE.COM"]
        self.end_marker = "STATEMENT OF OPEN INVOICE(S)"
        self.skip_lines = {"Statement Date:", "Total Due:", "www.unitedcorporate.com"}

#### Company Data Loading (O(1) Optimized)  
        self.dnm_companies, self.normalized_company_map = self._load_dnm_companies()
        
#### Extraction Comparison Tracking (ULTIMATE Feature)
        self.extraction_comparisons = []
        
#### Performance Optimizations
        self._processed_pages: Set[int] = set()
        gc.collect()

################################################################################













#### PATTERN CREATION AND NORMALIZATION
################################################################################

    def _create_business_suffix_pattern(self) -> re.Pattern:
        """Create comprehensive business suffix pattern for normalization."""
        suffixes = [
            'inc', 'incorporated', 'corp', 'corporation', 'llc', 'ltd', 'limited',
            'llp', 'lp', 'pc', 'pa', 'pllc', 'plc', 'co', 'company', 'companies',
            'enterprise', 'enterprises', 'group', 'groups', 'holding', 'holdings',
            'international', 'intl', 'global', 'solutions', 'services', 'systems',
            'technologies', 'tech', 'industries', 'foundation', 'trust', 'association',
            'society', 'institute', 'center', 'centre', 'organization', 'org'
        ]
        
        escaped_suffixes = [re.escape(suffix) for suffix in suffixes]
        pattern = r'\b(?:' + '|'.join(escaped_suffixes) + r')\b'
        return re.compile(pattern, re.IGNORECASE)

    def _normalize_company_name(self, name: str) -> str:
        """Normalize company names for consistent matching - O(1) operation."""
        if not name:
            return ""
        
        normalized = str(name).lower().strip()
        normalized = self.patterns['business_suffixes'].sub('', normalized)
        normalized = self.patterns['clean_text'].sub('', normalized)
        
        return normalized.strip()

################################################################################













#### COMPANY DATA LOADING AND MATCHING
################################################################################

    def _load_dnm_companies(self) -> Tuple[List[str], Dict[str, str]]:
        """Load DNM companies with O(1) lookup optimization."""
        try:
            workbook = load_workbook(self.excel_path, read_only=True)
            worksheet = workbook['10-2018']
            
            companies = []
            for row in worksheet.iter_rows(min_row=4, max_col=1, values_only=True):
                cell_value = row[0]
                if cell_value and str(cell_value).strip():
                    companies.append(str(cell_value).strip())
            
            workbook.close()
            
            # Create normalized mapping for O(1) lookups
            normalized_map = {}
            for company in companies:
                normalized = self._normalize_company_name(company)
                if normalized:
                    normalized_map[normalized] = company
            
            print(f"Loaded {len(companies)} companies from Excel")
            return companies, normalized_map
            
        except Exception as e:
            raise RuntimeError(f"Failed to load DNM companies: {e}")

    def _find_company_match(self, company_name: str) -> Tuple[Optional[str], List[Dict[str, str]]]:
        """ULTIMATE matching - Check ALL companies above 60% threshold."""
        # O(1) exact match check
        if company_name in self.dnm_companies:
            return company_name, []
        
        # O(1) normalized exact match
        normalized = self._normalize_company_name(company_name)
        if normalized in self.normalized_company_map:
            return self.normalized_company_map[normalized], []
        
        # ULTIMATE: Check ALL companies above 60% threshold (not just first match)
        similar_matches = []
        if normalized:
            for norm_company, original_company in self.normalized_company_map.items():
                similarity_score = SequenceMatcher(None, normalized, norm_company).ratio() * 100
                if similarity_score >= 60.0:
                    similar_matches.append({
                        "company_name": original_company,
                        "percentage": f"{round(similarity_score, 1)}%"
                    })
            
            # Sort by similarity score (highest first)
            similar_matches.sort(key=lambda x: float(x["percentage"].replace('%', '')), reverse=True)
        
        return None, similar_matches

################################################################################













#### ULTIMATE COMPANY NAME EXTRACTION (FIXED METHOD)
################################################################################

    def _extract_company_name_ultimate(self, text: str, lines: List[str], page_num: int, 
                                       current: int, total: int) -> Tuple[str, str, str]:
        """ULTIMATE company name extraction with all fixes from data.py."""
        # Get fallback method for comparison logging (top company name)
        fallback_company = lines[0].strip() if lines else "N/A"
        extraction_method = "unknown"
        
        # Method 1: Try to find company name between "Subtotal" and "Total Due" (most accurate)
        subtotal_match = self.patterns['total_due_from_subtotal'].search(text)
        if subtotal_match:
            company_raw = subtotal_match.group(1).strip()
            company = self.patterns['whitespace'].sub(' ', company_raw).strip()
            extraction_method = "subtotal_to_total_pattern"
        else:
            # Method 2: Try multiline pattern for company names split across lines
            multiline_match = self.patterns['total_due_multiline'].search(text)
            if multiline_match:
                company_raw = multiline_match.group(1).strip()
                company = self.patterns['whitespace'].sub(' ', company_raw.replace('\n', ' ')).strip()
                extraction_method = "multiline_pattern"
                
                # Validate: if it's reasonable length (< 100 chars), use it
                if len(company) <= 100:
                    pass  # Keep the multiline result
                else:
                    # Too long, try single line pattern instead
                    line_end_match = self.patterns['total_due_line_end'].search(text)
                    if line_end_match:
                        company_raw = line_end_match.group(1).strip()
                        company = self.patterns['whitespace'].sub(' ', company_raw).strip()
                        extraction_method = "line_end_pattern"
                        if len(company) > 100:
                            company = lines[0].strip()
                            extraction_method = "first_line_fallback"
                    else:
                        company = lines[0].strip()
                        extraction_method = "first_line_fallback"
            else:
                # Method 3: Look for company name at end of line before "Total Due $X.XX"
                line_end_match = self.patterns['total_due_line_end'].search(text)
                if line_end_match:
                    company_raw = line_end_match.group(1).strip()
                    company = self.patterns['whitespace'].sub(' ', company_raw).strip()
                    extraction_method = "line_end_pattern"
                    
                    # Validate: if it's too long (> 100 chars), it likely captured invoice data
                    if len(company) > 100:
                        # Use first line instead
                        company = lines[0].strip()
                        extraction_method = "first_line_fallback"
                else:
                    # Method 4: Last resort - use first line from cleaned content
                    company = lines[0].strip()
                    extraction_method = "first_line_fallback"
        
        # Log comparison for ULTIMATE tracking
        if company.strip() != fallback_company.strip():
            self.extraction_comparisons.append({
                'page_num': page_num,
                'current_page': current,
                'total_pages': total,
                'old_method': fallback_company,
                'new_method': company,
                'extraction_method': extraction_method,
                'match': company.strip() == fallback_company.strip()
            })
        
        return company, extraction_method, fallback_company

################################################################################













#### LOCATION AND DESTINATION LOGIC
################################################################################

    def _detect_location(self, text: str) -> str:
        """Detect location using optimized state detection - O(k) where k = 50 states."""
        text_upper = f" {text.upper()} "
        return "National" if any(f" {state} " in text_upper for state in self.us_states) else "Foreign"

    def _determine_destination_enhanced(self, exact_match: Optional[str], text: str, location: str, 
                            pages: int, best_percentage: float, has_email: bool) -> str:
        """Enhanced destination logic with improved accuracy."""
        if exact_match:
            return "DNM"
        if has_email:
            return "DNM"
        if best_percentage >= 90.0:
            return "DNM"
        
        if location == "Foreign":
            return "Foreign"
        return "Natio Single" if pages == 1 else "Natio Multi"

################################################################################













#### ULTIMATE STATEMENT EXTRACTION
################################################################################

    def _extract_statement_data(self, text: str, page_num: int) -> Optional[Dict[str, Any]]:
        """ULTIMATE statement extraction combining all improvements."""
        # Parse page information
        page_match = self.patterns['page'].search(text)
        current_page, total_pages = (int(page_match.group(1)), int(page_match.group(2))) if page_match else (1, 1)
        
        # Find content boundaries
        start_pos = min((text.find(marker) for marker in self.start_markers if marker in text), default=-1)
        end_pos = text.find(self.end_marker)
        
        if start_pos == -1 or end_pos == -1 or start_pos >= end_pos:
            return None
        
        # Extract and clean content
        content = text[start_pos:end_pos]
        for marker in self.start_markers:
            content = content.replace(marker, '')
        
        lines = [
            line.strip() for line in content.splitlines()
            if line.strip() and not any(skip in line for skip in self.skip_lines)
        ]
        
        if not lines:
            return None

#### ULTIMATE Company Name Extraction
        company_name, extraction_method, top_company_name = self._extract_company_name_ultimate(
            text, lines, page_num, current_page, total_pages
        )

#### Additional Data Processing
        rest_text = "\n".join(lines[1:])
        location = self._detect_location(rest_text)
        exact_match, similar_matches = self._find_company_match(company_name)

#### Calculate Page Range (O(1) Operation)
        if total_pages == 1:
            page_range = str(page_num)
            first_page = page_num
        else:
            start_page = page_num - (current_page - 1)
            page_range = "-".join(map(str, range(start_page, start_page + total_pages)))
            first_page = start_page

#### Determine Processing Flags
        has_email = "email" in rest_text.lower()
        best_match = similar_matches[0] if similar_matches else None
        best_percentage = float(best_match["percentage"].replace('%', '')) if best_match else 0
        is_high_confidence = best_percentage >= 90.0
        
        manual_required = False
        ask_question = False
        
        if not (has_email or is_high_confidence or exact_match):
            manual_required = len(similar_matches) > 0
            if manual_required:
                ask_question = best_percentage < 90.0

#### Determine Destination
        destination = self._determine_destination_enhanced(
            exact_match, rest_text, location, total_pages, best_percentage, has_email
        )

        # Build the result dictionary with proper field ordering
        result = OrderedDict()
        result["company_name"] = company_name
        
        # ONLY add unusedCompanyName when bottom company name != top company name
        if company_name.strip() != top_company_name.strip():
            result["unusedCompanyName"] = top_company_name
            
        result["exact_match"] = exact_match
        result["similar_matches"] = similar_matches  # ALL matches above 60%
        result["manual_required"] = manual_required
        result["ask_question"] = ask_question
        result["rest_of_lines"] = rest_text
        result["location"] = location
        result["paging"] = f"page {current_page} of {total_pages}"
        result["number_of_pages"] = str(total_pages)
        result["page_number_in_uploaded_pdf"] = page_range
        result["first_page_number"] = first_page
        result["destination"] = destination
        result["extraction_method"] = extraction_method  # ULTIMATE: Track extraction method
            
        return result

################################################################################













#### ULTIMATE EXTRACTION WORKFLOW
################################################################################

    def extract_statements(self) -> List[Dict[str, Any]]:
        """ULTIMATE extraction workflow with optimized processing."""
        try:
            doc = fitz.open(str(self.pdf_path))
            statements = []
            
            print(f"Processing {len(doc)} pages with {len(self.dnm_companies)} DNM companies loaded...")
            
            for page_idx in range(len(doc)):
                page_num = page_idx + 1
                
                if page_num in self._processed_pages:
                    continue
                
                statement_data = self._extract_statement_data(doc.load_page(page_idx).get_text(), page_num)
                
                if statement_data:
                    statements.append(statement_data)
                    print(f" Found: {statement_data['company_name']}")
                    
                    # Mark pages as processed to avoid reprocessing
                    total_pages = int(statement_data["number_of_pages"])
                    if total_pages > 1:
                        page_range = statement_data["page_number_in_uploaded_pdf"].split("-")
                        self._processed_pages.update(range(int(page_range[0]), int(page_range[-1]) + 1))
                    else:
                        self._processed_pages.add(page_num)
            
            doc.close()
            return statements
            
        except Exception as e:
            raise RuntimeError(f"Failed to extract statements: {e}")

################################################################################













#### ULTIMATE JSON OUTPUT WITH INTEGRATED LOGGING
################################################################################

    def save_results(self, statements: List[Dict[str, Any]], output_path: Optional[str] = None) -> str:
        """ULTIMATE save with integrated extraction comparison logging."""
        if not output_path:
            today = datetime.now().strftime("%b%d%Y").lower()
            output_path = f"{today}.json"
            
            counter = 1
            while os.path.exists(output_path):
                output_path = f"{today}-{counter}.json"
                counter += 1

#### ULTIMATE Data Structure with Extraction Comparison
        data = {
            "dnm_companies": self.dnm_companies,
            "extracted_statements": statements,
            "total_statements_found": len(statements),
            "processing_timestamp": datetime.now().isoformat(),
            
            # ULTIMATE: Integrated extraction comparison logging
            "extraction_comparison_log": {
                "total_statements_with_different_extractions": len(self.extraction_comparisons),
                "extraction_details": self.extraction_comparisons,
                "summary": {
                    "extraction_methods_used": list(set(comp['extraction_method'] for comp in self.extraction_comparisons)),
                    "pages_with_multiline_extraction": len([c for c in self.extraction_comparisons if c['extraction_method'] == 'multiline_pattern']),
                    "pages_with_subtotal_extraction": len([c for c in self.extraction_comparisons if c['extraction_method'] == 'subtotal_to_total_pattern']),
                    "pages_with_improved_accuracy": len([c for c in self.extraction_comparisons if c['extraction_method'] != 'first_line_fallback'])
                }
            }
        }
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            
            print(f" Results saved to {output_path}")
            return output_path
            
        except Exception as e:
            raise RuntimeError(f"Failed to save results: {e}")

################################################################################













#### PRODUCTION WORKFLOW METHODS
################################################################################

    def process_interactive_questions(self, statements: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Process interactive questions for companies requiring manual review."""
        questions_needed = [stmt for stmt in statements if stmt.get('ask_question', False)]
        
        if not questions_needed:
            print(" No manual questions required.")
            return statements
        
        print(f"\nFound {len(questions_needed)} companies requiring manual review:")
        
        for i, statement in enumerate(questions_needed):
            company_name = statement.get('company_name', 'Unknown')
            similar_matches = statement.get('similar_matches', [])
            
            if similar_matches:
                best_match = similar_matches[0]
                print(f"\nQuestion {i + 1} of {len(questions_needed)}:")
                print(f"Company '{company_name}' is similar to '{best_match['company_name']}' ({best_match['percentage']})")
                print("Are they the same company? (y/n/s to skip all)")
                
                while True:
                    try:
                        response = input("> ").strip().lower()
                        
                        if response == 'y':
                            statement['destination'] = 'DNM'
                            statement['user_answered'] = 'yes'
                            print(f" Marked '{company_name}' as DNM")
                            break
                        elif response == 'n':
                            statement['user_answered'] = 'no'
                            print(f" Kept '{company_name}' as {statement['destination']}")
                            break
                        elif response == 's':
                            print(" Skipping remaining questions")
                            return statements
                        else:
                            print("Please enter 'y', 'n', or 's'")
                            
                    except (KeyboardInterrupt, EOFError):
                        print("\nOperation cancelled.")
                        sys.exit(0)
        
        return statements

    def create_split_pdfs(self, statements: List[Dict[str, Any]]) -> Dict[str, int]:
        """Create destination-based PDF splits - O(n) operation."""
        destinations = {"DNM": [], "Foreign": [], "Natio Single": [], "Natio Multi": []}
        
        for statement in statements:
            dest = statement.get('destination', '').strip()
            if dest in destinations:
                destinations[dest].append(statement)
        
        output_files = {
            "DNM": "DNM.pdf",
            "Foreign": "Foreign.pdf", 
            "Natio Single": "natioSingle.pdf",
            "Natio Multi": "natioMulti.pdf"
        }
        
        results = {}
        
        try:
            reader = PdfReader(str(self.pdf_path))
            total_pages = len(reader.pages)
            
            for dest, statements_list in destinations.items():
                if not statements_list:
                    continue
                
                writer = PdfWriter()
                pages_added = 0
                
                for statement in statements_list:
                    page_range = statement.get('page_number_in_uploaded_pdf', '')
                    for page_str in page_range.split('-'):
                        try:
                            page_num = int(page_str.strip()) - 1
                            if 0 <= page_num < total_pages:
                                writer.add_page(reader.pages[page_num])
                                pages_added += 1
                        except ValueError:
                            continue
                
                if pages_added > 0:
                    output_path = output_files[dest]
                    with open(output_path, 'wb') as f:
                        writer.write(f)
                    results[dest] = pages_added
                    print(f" Created {output_path} with {pages_added} pages")
            
            return results
            
        except Exception as e:
            raise RuntimeError(f"Failed to create split PDFs: {e}")

################################################################################













#### ULTIMATE COMPLETE WORKFLOW
################################################################################

    def run_complete_workflow(self, skip_questions: bool = False) -> bool:
        """ULTIMATE workflow execution."""
        try:
            print("=" * 80)
            print("               ULTIMATE STATEMENT PROCESSOR")
            print("=" * 80)
            print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print()

#### Step 1: Extract Statements
            print(" Step 1: Extracting statements from PDF...")
            statements = self.extract_statements()
            print(f" Extracted {len(statements)} statements")

#### Step 2: Process Interactive Questions
            if not skip_questions:
                print("\n Step 2: Processing manual questions...")
                statements = self.process_interactive_questions(statements)
                print(" Manual questions processed")
            else:
                print("\n⏭️  Step 2: Skipping interactive questions...")

#### Step 3: Save Results with Integrated Logging
            print("\n Step 3: Saving results with extraction comparison...")
            output_file = self.save_results(statements)
            print(" Results saved with integrated logging")

#### Step 4: Create Split PDFs
            if not skip_questions:
                print("\n Step 4: Creating destination PDFs...")
                split_results = self.create_split_pdfs(statements)
                print(" PDFs created successfully")

#### Final Summary
                print("\n" + "=" * 80)
                print(" ULTIMATE WORKFLOW COMPLETED SUCCESSFULLY!")
                print("=" * 80)
                
                print(f"Total statements processed: {len(statements)}")
                print(f"Extraction comparisons logged: {len(self.extraction_comparisons)}")
                print(f"JSON output: {output_file}")
                print("PDF outputs created:")
                for dest, pages in split_results.items():
                    print(f"  • {dest}: {pages} pages")
            else:
                print("\n" + "=" * 80)
                print(" ULTIMATE EXTRACTION COMPLETED")
                print("=" * 80)
                print(f"Total statements processed: {len(statements)}")
                print(f"JSON output: {output_file}")
            
            print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            return True
            
        except Exception as e:
            print(f"\n Workflow failed: {e}")
            return False


################################################################################













# UTILITY FUNCTIONS
################################################################################

def find_files_in_directory() -> Tuple[Optional[str], Optional[str]]:
    """Auto-detect PDF and Excel files in current directory."""
    pdf_files = list(Path('.').glob('*.pdf'))
    excel_files = list(Path('.').glob('*.xlsx')) + list(Path('.').glob('*.xls'))
    
    if pdf_files and excel_files:
        return str(pdf_files[0]), str(excel_files[0])
    return None, None


def get_file_paths() -> Tuple[str, str]:
    """Get file paths from environment or use auto-detection."""
    # Try environment variables first
    pdf_path = os.environ.get('PDF_PATH')
    excel_path = os.environ.get('EXCEL_PATH')
    
    if pdf_path and excel_path:
        print(f"Using environment variables:")
        print(f"  PDF_PATH: {pdf_path}")
        print(f"  EXCEL_PATH: {excel_path}")
        return pdf_path, excel_path
    
    # Auto-detect files in current directory
    detected_pdf, detected_excel = find_files_in_directory()
    if detected_pdf and detected_excel:
        print(f"Auto-detected files:")
        print(f"  PDF: {detected_pdf}")
        print(f"  Excel: {detected_excel}")
        return detected_pdf, detected_excel
    
    raise FileNotFoundError("No PDF or Excel files found. Set PDF_PATH and EXCEL_PATH environment variables.")


# LOCAL TESTING FUNCTION - COMMENT OUT FOR PRODUCTION
def local_test():
    """Local testing function - COMMENT OUT FOR PRODUCTION"""
    pdf_path = "/Users/personal/Desktop/JSONExtract/UCS Statements 08-01-24.pdf"
    excel_path = "/Users/personal/Desktop/JSONExtract/Do Not Mail List For Statements.xlsx"
    
    processor = StatementProcessor(pdf_path, excel_path)
    statements = processor.extract_statements()
    output_path = processor.save_results(statements, "local_test_output.json")
    
    print(f" Local test completed - results saved to {output_path}")
    return output_path


################################################################################













# MAIN EXECUTION
################################################################################

def main() -> int:
    """ULTIMATE main entry point."""
    try:
        print(" ULTIMATE Statement Processing System")
        print("=" * 50)
        
        pdf_path, excel_path = get_file_paths()
        
        processor = StatementProcessor(pdf_path, excel_path)
        skip_questions = '--skip-questions' in sys.argv
        success = processor.run_complete_workflow(skip_questions)
        
        return 0 if success else 1
        
    except KeyboardInterrupt:
        print("\n\n⏹️  Process interrupted by user")
        return 1
    except Exception as e:
        print(f"\n Fatal error: {e}")
        return 1


if __name__ == "__main__":
    # UNCOMMENT BELOW FOR LOCAL TESTING - COMMENT OUT FOR PRODUCTION
    # local_test()
    
    sys.exit(main())